# -*- coding: utf-8 -*-
"""
T.C. Sanayi ve Teknoloji Bakanlığı - Kocaeli İl Müdürlüğü
İdari Para Cezaları (İPC) Yönetim Sistemi — BACKEND
Flask + SQLite  |  Frontend: ../frontend (templates + static)

Contributor credits: see ../AUTHORS.md
"""
import os, re, json, sqlite3, io, tempfile, time, uuid, mimetypes, sys, secrets
from datetime import datetime, date, timedelta
from functools import wraps

from flask import (Flask, render_template, request, redirect, url_for,
                   session, flash, g, send_file, send_from_directory, abort)
from werkzeug.security import generate_password_hash, check_password_hash
from itsdangerous import URLSafeTimedSerializer, BadSignature, SignatureExpired
from importer import (
    IMPORT_FIELD_DEFINITIONS,
    ImportWorkbookError,
    file_sha256,
    instrument_summary,
    instrument_total,
    normalize_instrument_items,
    parse_workbook,
)

FROZEN_WINDOWS_APP = bool(getattr(sys, "frozen", False))
if FROZEN_WINDOWS_APP:
    RESOURCE_ROOT = getattr(
        sys, "_MEIPASS", os.path.dirname(os.path.abspath(sys.executable))
    )
    BASE_DIR = os.path.join(RESOURCE_ROOT, "backend")
    FRONTEND = os.path.join(RESOURCE_ROOT, "frontend")
    DEFAULT_DATA_DIR = os.path.join(
        os.path.dirname(os.path.abspath(sys.executable)), "data"
    )
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    FRONTEND = os.path.join(os.path.dirname(BASE_DIR), "frontend")
    DEFAULT_DATA_DIR = BASE_DIR

DATA_DIR = os.environ.get("IPC_DATA_DIR", DEFAULT_DATA_DIR)
DB_PATH = os.environ.get("IPC_DB_PATH", os.path.join(DATA_DIR, "ipc.db"))
IMPORT_TMP_DIR = os.environ.get(
    "IPC_IMPORT_TMP_DIR", os.path.join(tempfile.gettempdir(), "ipc-imports")
)
BACKUP_DIR = os.environ.get("IPC_BACKUP_DIR", os.path.join(DATA_DIR, "backups"))
UPLOAD_DIR = os.environ.get("IPC_UPLOAD_DIR", os.path.join(DATA_DIR, "uploads"))

MAX_ATTACHMENT_COUNT = 10
MAX_ATTACHMENT_SIZE = 100 * 1024 * 1024
ALLOWED_ATTACHMENT_EXTENSIONS = {
    ".pdf", ".txt", ".csv", ".zip", ".rar",
    ".doc", ".docx", ".xls", ".xlsx", ".ppt", ".pptx",
    ".jpg", ".jpeg", ".png", ".webp", ".gif", ".heic",
    ".mp4", ".webm", ".mov", ".avi", ".mkv",
}
INLINE_ATTACHMENT_EXTENSIONS = {
    ".pdf", ".jpg", ".jpeg", ".png", ".webp", ".gif", ".mp4", ".webm",
}

PRODUCTION = os.environ.get("IPC_ENV", "development").strip().lower() == "production"
SECURE_COOKIES = os.environ.get(
    "IPC_SECURE_COOKIES", "1" if PRODUCTION else "0"
).strip().lower() in ("1", "true", "yes")
SECRET_KEY = os.environ.get("IPC_SECRET_KEY")
if PRODUCTION and not SECRET_KEY:
    raise RuntimeError(
        "Production requires the IPC_SECRET_KEY environment variable."
    )

INITIAL_ADMIN_NAME = os.environ.get("IPC_INITIAL_ADMIN_NAME", "admin").strip()
INITIAL_ADMIN_EMAIL = os.environ.get(
    "IPC_INITIAL_ADMIN_EMAIL", "admin@example.local"
).strip()
INITIAL_ADMIN_PASSWORD = os.environ.get("IPC_INITIAL_ADMIN_PASSWORD")

app = Flask(__name__,
            template_folder=os.path.join(FRONTEND, "templates"),
            static_folder=os.path.join(FRONTEND, "static"))
app.config.update(
    SECRET_KEY=SECRET_KEY or secrets.token_urlsafe(48),
    MAX_CONTENT_LENGTH=300 * 1024 * 1024,
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=SECURE_COOKIES,
)
serializer = URLSafeTimedSerializer(app.config["SECRET_KEY"])

@app.errorhandler(413)
def upload_too_large(_error):
    flash(
        "Yüklenen dosyaların toplam boyutu 300 MB sınırını aşıyor.",
        "danger",
    )
    return redirect(url_for("sirket_ekle"))

DATATYPES = [
    "INT", "VARCHAR(n)", "CHAR(n)", "TEXT", "DATETIME / TIMESTAMP",
    "DECIMAL", "FLOAT / DOUBLE", "BOOLEAN", "TARİH (GG.AA.YYYY)", "AÇILIR LİSTE"
]
PARA_TIPLERI = ("DECIMAL", "FLOAT / DOUBLE")

# ---------------------------------------------------------------- veritabanı
def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH, timeout=30)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA busy_timeout=30000")
    return g.db

@app.teardown_appcontext
def close_db(exc):
    db = g.pop("db", None)
    if db is not None:
        db.close()

DEFAULT_FIELDS = [
    ("sira_no", "Sıra No", "INT", 1, None, 1),
    ("defter_sira_no", "Sorumlu Personel", "VARCHAR(n)", 0, None, 1),
    ("yil", "Yılı", "INT", 1, None, 0),
    ("il_adi", "İlin Adı (Sorumlu Teşkilat)", "VARCHAR(n)", 1, None, 0),
    ("hukuki_dayanak", "HUKUKİ DAYANAK", "AÇILIR LİSTE", 1,
     ["3516.15/c", "3516.15/e", "3516.15/h", "3516.15/i"], 0),
    ("olcu_aleti_sayisi", "ÖLÇÜ ALETİ SAYISI", "INT", 1, None, 0),
    ("olcu_aleti_cinsi", "ÖLÇÜ ALETİ CİNSİ", "AÇILIR LİSTE", 1,
     ["2. SINIF OTOMATİK OLMAYAN TARTI ALETİ",
      "3. SINIF OTOMATİK OLMAYAN TARTI ALETİ",
      "AdBlue Sayacı", "AddBlue Sayacı", "Akaryakıt ve LPG Sayacı",
      "Egzoz Emisyon Cihazı", "Kanuna Uygun Olmayan Ölçü Aleti",
      "Lastik Hava Basınç Ölçer", "Mevzuata Aykırı Faaliyette Bulunmak",
      "Tanker sayacı", "Otomatik Olmayan 2T. Üzeri Tartı A.",
      "Taksimetre"], 0),
    ("kaynak_notu", "Kaynak Notu / Uyarısı (başlıksız H sütunu)", "TEXT", 0, None, 0),
    ("tablet_tutanak_no", "TABLET TUTANAK NUMARASI", "VARCHAR(n)", 1, None, 0),
    ("cezanin_muhatabi", "Cezanın Muhatabı (Kişi/Firma Adı)", "VARCHAR(n)", 1, None, 0),
    ("ceza_onay_tarihi", "Cezanın Onay Tarihi (Kararın Verildiği Tarih, İYK Tarihi)", "TARİH (GG.AA.YYYY)", 1, None, 0),
    ("ceza_tutari", "Verilen Cezanın Tutarı (TL)", "DECIMAL", 1, None, 0),
    ("gonderim_turu", "GÖNDERİM TÜRÜ", "AÇILIR LİSTE", 1,
     ["ELDEN", "FİZİKİ", "UETS", "İPTAL EDİLDİ"], 0),
    ("teblig_tarihi", "Cezanın İlgiliye Tebliğ Tarihi", "TARİH (GG.AA.YYYY)", 1, None, 0),
    ("dogrudan_yatirilan", "Tebligat Üzerine İlgili Tarafından Doğrudan Yatırılan Tutar (TL)", "DECIMAL", 0, None, 0),
    ("odeme_tarihi", "ödeme tarihi", "TARİH (GG.AA.YYYY)", 0, None, 0),
    ("kesinlesme_tarihi", "İtiraz Edilmeyenlerde Kesinleşme Tarihi", "TARİH (GG.AA.YYYY)", 0, None, 0),
    ("kalan_sure", "KALAN SÜRE", "INT", 0, None, 1),
    ("vergi_bildirim_tarihi", "İtiraz Edilmeyenlerde Tahsilat İçin İlgili Vergi Dairesine Bildirim Tarihi", "TARİH (GG.AA.YYYY)", 0, None, 0),
    ("vergi_dairesi", "Bildirim Yapılan Vergi Dairesinin Adı", "AÇILIR LİSTE", 0,
     ["İlyasbey VD.", "Körfez VD.", "Merter VD.", "Orhangazi VD.",
      "Tepecik VD.", "Uluçınar VD", "Yakacık VD", "Yenikapı VD."], 0),
    ("itiraz_sonucu", "İtirazın Sonucu (Kurum Lehine/ Kurum Aleyhine)", "TEXT", 0, None, 0),
    ("itiraz_lehine_tutar", "İtiraz Üzerine Kurum Lehine Sonuçlananlarda İlgili Tarafından Doğrudan Yatırılan Tutar (TL)", "DECIMAL", 0, None, 0),
    ("itiraz_vergi_bildirim_tarihi", "İtiraz Üzerine Kurum Lehine Sonuçlananlarda Tahsilat İçin İlgili Vergi Dairesine Bildirim Tarihi", "TARİH (GG.AA.YYYY)", 0, None, 0),
    ("vergi_tahsil_tutari", "İlgili Vergi Dairesince Tahsil Edildiği Bildirilen Tutar (TL)", "DECIMAL", 0, None, 0),
    ("vergi_tahsil_tarihi", "İlgili Vergi Dairesince Tahsil Edildiği Bildirilen Tarih", "TARİH (GG.AA.YYYY)", 0, None, 0),
    ("tespit_kurum", "Tespiti Yapan Kurum", "AÇILIR LİSTE", 1,
     ["Damdacı", "Gebze Kalibrasyon", "İL MÜD.", "Safir Servis", "TSE"], 0),
    ("ebys", "Tespit Yazısının EBYS Tarih ve Sayısı", "VARCHAR(n)", 1, None, 0),
    ("kaynak_durumu", "DURUMU", "TEXT", 0, None, 0),
]

REAL_EXCEL_COLUMNS = {
    "sira_no": "A",
    "defter_sira_no": "B",
    "yil": "C",
    "il_adi": "D",
    "hukuki_dayanak": "E",
    "olcu_aleti_sayisi": "F",
    "olcu_aleti_cinsi": "G",
    "kaynak_notu": "H",
    "tablet_tutanak_no": "I",
    "cezanin_muhatabi": "J",
    "ceza_onay_tarihi": "K",
    "ceza_tutari": "L",
    "gonderim_turu": "M",
    "teblig_tarihi": "N",
    "dogrudan_yatirilan": "O",
    "odeme_tarihi": "P",
    "kesinlesme_tarihi": "Q",
    "kalan_sure": "R",
    "vergi_bildirim_tarihi": "S",
    "vergi_dairesi": "T",
    "itiraz_sonucu": "U",
    "itiraz_lehine_tutar": "V",
    "itiraz_vergi_bildirim_tarihi": "W",
    "vergi_tahsil_tutari": "X",
    "vergi_tahsil_tarihi": "Y",
    "tespit_kurum": "Z",
    "ebys": "AA",
    "kaynak_durumu": "AB",
}

DETAIL_SECTION_KEYS = [
    (
        "Tespit ve Denetim",
        "Kaydın kaynağı, denetim bilgileri ve sorumlu personel.",
        (
            "defter_sira_no", "yil", "il_adi", "hukuki_dayanak",
            "olcu_aleti_sayisi", "olcu_aleti_cinsi", "tespit_kurum",
            "ebys", "kaynak_notu",
        ),
    ),
    (
        "Ceza Kararı",
        "Karar, muhatap ve uygulanan idari para cezası.",
        (
            "tablet_tutanak_no", "cezanin_muhatabi", "ceza_onay_tarihi",
            "ceza_tutari", "kaynak_durumu",
        ),
    ),
    (
        "Tebligat ve Ödeme",
        "Tebligat süreci, kalan süre ve doğrudan ödeme bilgileri.",
        (
            "gonderim_turu", "teblig_tarihi", "kalan_sure",
            "dogrudan_yatirilan", "odeme_tarihi",
        ),
    ),
    (
        "Kesinleşme ve Tahsilat",
        "Kesinleşen kayıtların vergi dairesi bildirim ve tahsilat bilgileri.",
        (
            "kesinlesme_tarihi", "vergi_bildirim_tarihi", "vergi_dairesi",
            "vergi_tahsil_tutari", "vergi_tahsil_tarihi",
        ),
    ),
    (
        "İtiraz",
        "İtiraz sonucu ve itiraz sonrasındaki tahsilat süreci.",
        (
            "itiraz_sonucu", "itiraz_lehine_tutar",
            "itiraz_vergi_bildirim_tarihi",
        ),
    ),
]

REPORT_PRESETS = {
    "ozet": {
        "title": "Yönetim Özeti",
        "description": "Temel kayıt, tutar, tebligat ve durum bilgileri.",
        "keys": (
            "sira_no", "yil", "tablet_tutanak_no", "cezanin_muhatabi",
            "hukuki_dayanak", "ceza_tutari", "teblig_tarihi", "kalan_sure",
        ),
    },
    "tam": {
        "title": "Tam Kayıt Raporu",
        "description": "Tüm İPC alanlarını içerir; PDF'de her kayıt okunabilir ayrıntı düzeninde gösterilir.",
        "keys": tuple(REAL_EXCEL_COLUMNS),
    },
    "tahsilat": {
        "title": "Tahsilat Raporu",
        "description": "Ödeme, kesinleşme ve vergi dairesi sürecine odaklanır.",
        "keys": (
            "sira_no", "yil", "tablet_tutanak_no", "cezanin_muhatabi",
            "ceza_tutari", "dogrudan_yatirilan", "odeme_tarihi",
            "kesinlesme_tarihi", "vergi_bildirim_tarihi", "vergi_dairesi",
            "vergi_tahsil_tutari", "vergi_tahsil_tarihi",
        ),
    },
    "itiraz": {
        "title": "İtiraz Raporu",
        "description": "İtiraz sonucu ve itiraz sonrası tahsilat bilgileri.",
        "keys": (
            "sira_no", "yil", "tablet_tutanak_no", "cezanin_muhatabi",
            "ceza_tutari", "itiraz_sonucu", "itiraz_lehine_tutar",
            "itiraz_vergi_bildirim_tarihi", "vergi_dairesi",
        ),
    },
}

STATUS_FILTER_LABELS = {
    "yeni": "Yeni kayıt",
    "teblig-bekliyor": "Tebliğ bekliyor",
    "sure-isliyor": "Süresi işleyen",
    "sure-doldu": "Süresi dolan",
    "kesinlesti": "Kesinleşen",
    "vergi-bildirildi": "Vergi dairesine bildirilen",
    "itiraz": "İtiraz sürecindeki",
    "odendi": "Ödenen",
    "iptal": "İptal edilen",
}

def init_db():
    database_parent = os.path.dirname(os.path.abspath(DB_PATH))
    os.makedirs(database_parent, exist_ok=True)
    db = sqlite3.connect(DB_PATH, timeout=30)
    db.execute("PRAGMA journal_mode=WAL")
    db.execute("PRAGMA busy_timeout=30000")
    db.execute("PRAGMA synchronous=NORMAL")
    c = db.cursor()
    c.executescript("""
    CREATE TABLE IF NOT EXISTS users(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        email TEXT NOT NULL UNIQUE,
        password_hash TEXT NOT NULL,
        role TEXT NOT NULL DEFAULT 'personel',
        created_at TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS form_fields(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        field_key TEXT NOT NULL UNIQUE,
        label TEXT NOT NULL,
        datatype TEXT NOT NULL,
        required INTEGER DEFAULT 0,
        options TEXT,
        is_system INTEGER DEFAULT 0,
        position INTEGER DEFAULT 0
    );
    CREATE TABLE IF NOT EXISTS companies(
        sira_no INTEGER PRIMARY KEY AUTOINCREMENT,
        data TEXT NOT NULL,
        created_by TEXT,
        created_at TEXT DEFAULT (datetime('now','localtime')),
        updated_at TEXT
    );
    CREATE TABLE IF NOT EXISTS pending_actions(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        action_type TEXT NOT NULL,
        sira_no INTEGER,
        data TEXT,
        requested_by INTEGER NOT NULL,
        status TEXT DEFAULT 'bekliyor',
        created_at TEXT DEFAULT (datetime('now','localtime')),
        decided_at TEXT,
        decision_reason TEXT
    );
    CREATE TABLE IF NOT EXISTS notifications(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        title TEXT,
        message TEXT NOT NULL,
        kind TEXT NOT NULL DEFAULT 'general',
        priority TEXT NOT NULL DEFAULT 'info',
        action_url TEXT,
        dedupe_key TEXT,
        metadata TEXT,
        is_read INTEGER DEFAULT 0,
        resolved_at TEXT,
        created_at TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS import_history(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        filename TEXT NOT NULL,
        file_sha256 TEXT NOT NULL,
        sheet_name TEXT,
        source_rows INTEGER NOT NULL,
        inserted_rows INTEGER NOT NULL,
        updated_rows INTEGER NOT NULL,
        skipped_rows INTEGER NOT NULL,
        warning_count INTEGER NOT NULL,
        conflict_mode TEXT NOT NULL,
        backup_path TEXT,
        imported_by INTEGER,
        imported_at TEXT DEFAULT (datetime('now','localtime'))
    );
    CREATE TABLE IF NOT EXISTS app_meta(
        key TEXT PRIMARY KEY,
        value TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS attachments(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_sira_no INTEGER,
        pending_action_id INTEGER,
        original_name TEXT NOT NULL,
        stored_name TEXT NOT NULL UNIQUE,
        mime_type TEXT,
        file_size INTEGER NOT NULL,
        uploaded_by INTEGER,
        created_at TEXT DEFAULT (datetime('now','localtime')),
        CHECK(company_sira_no IS NOT NULL OR pending_action_id IS NOT NULL)
    );
    CREATE INDEX IF NOT EXISTS idx_attachments_company
        ON attachments(company_sira_no);
    CREATE INDEX IF NOT EXISTS idx_attachments_pending
        ON attachments(pending_action_id);
    """)
    company_columns = {
        row[1] for row in c.execute("PRAGMA table_info(companies)").fetchall()
    }
    if "source_file_hash" not in company_columns:
        c.execute("ALTER TABLE companies ADD COLUMN source_file_hash TEXT")
    if "source_row" not in company_columns:
        c.execute("ALTER TABLE companies ADD COLUMN source_row INTEGER")
    notification_columns = {
        row[1] for row in c.execute("PRAGMA table_info(notifications)").fetchall()
    }
    notification_migrations = {
        "title": "TEXT",
        "kind": "TEXT NOT NULL DEFAULT 'general'",
        "priority": "TEXT NOT NULL DEFAULT 'info'",
        "action_url": "TEXT",
        "dedupe_key": "TEXT",
        "metadata": "TEXT",
        "resolved_at": "TEXT",
    }
    for column, definition in notification_migrations.items():
        if column not in notification_columns:
            c.execute(f"ALTER TABLE notifications ADD COLUMN {column} {definition}")
    pending_columns = {
        row[1] for row in c.execute("PRAGMA table_info(pending_actions)").fetchall()
    }
    if "decision_reason" not in pending_columns:
        c.execute("ALTER TABLE pending_actions ADD COLUMN decision_reason TEXT")
    c.execute(
        """CREATE UNIQUE INDEX IF NOT EXISTS idx_notifications_active_dedupe
           ON notifications(user_id, dedupe_key)
           WHERE dedupe_key IS NOT NULL AND resolved_at IS NULL"""
    )
    if not c.execute("SELECT 1 FROM users LIMIT 1").fetchone():
        admin_password = INITIAL_ADMIN_PASSWORD
        if not admin_password:
            db.close()
            raise RuntimeError(
                "The first start requires IPC_INITIAL_ADMIN_PASSWORD. "
                "The Windows launcher asks for it automatically."
            )
        c.execute("INSERT INTO users(name,email,password_hash,role) VALUES(?,?,?,?)",
                  (INITIAL_ADMIN_NAME, INITIAL_ADMIN_EMAIL,
                   generate_password_hash(admin_password), "admin"))
    if not c.execute("SELECT 1 FROM form_fields LIMIT 1").fetchone():
        for i, (k, l, dt, req, opts, sys_) in enumerate(DEFAULT_FIELDS):
            c.execute("""INSERT INTO form_fields
                (field_key,label,datatype,required,options,is_system,position)
                VALUES(?,?,?,?,?,?,?)""",
                (k, l, dt, req, json.dumps(opts, ensure_ascii=False) if opts else None, sys_, i))

    schema_version = c.execute(
        "SELECT value FROM app_meta WHERE key='real_workbook_schema_version'"
    ).fetchone()
    if not schema_version or schema_version[0] != "1":
        canonical_keys = {field[0] for field in DEFAULT_FIELDS}
        c.execute("DELETE FROM form_fields WHERE field_key='kaynak_kalan_sure'")
        for position, (key, label, datatype, required, options, is_system) in enumerate(DEFAULT_FIELDS):
            existing = c.execute(
                "SELECT options FROM form_fields WHERE field_key=?", (key,)
            ).fetchone()
            if existing:
                current_options = json.loads(existing[0]) if existing[0] else []
                merged_options = list(current_options)
                for option in options or []:
                    if option not in merged_options:
                        merged_options.append(option)
                c.execute(
                    """UPDATE form_fields
                       SET label=?, position=?, is_system=?, options=?
                       WHERE field_key=?""",
                    (
                        label,
                        position,
                        is_system,
                        json.dumps(merged_options, ensure_ascii=False) if merged_options else None,
                        key,
                    ),
                )
            else:
                c.execute(
                    """INSERT INTO form_fields
                       (field_key,label,datatype,required,options,is_system,position)
                       VALUES(?,?,?,?,?,?,?)""",
                    (
                        key,
                        label,
                        datatype,
                        required,
                        json.dumps(options, ensure_ascii=False) if options else None,
                        is_system,
                        position,
                    ),
                )
        custom_fields = c.execute(
            "SELECT id, field_key FROM form_fields ORDER BY position, id"
        ).fetchall()
        custom_position = len(DEFAULT_FIELDS)
        for field_id, field_key in custom_fields:
            if field_key in canonical_keys:
                continue
            c.execute(
                "UPDATE form_fields SET position=? WHERE id=?",
                (custom_position, field_id),
            )
            custom_position += 1
        c.execute(
            """INSERT INTO app_meta(key,value)
               VALUES('real_workbook_schema_version','1')
               ON CONFLICT(key) DO UPDATE SET value=excluded.value"""
        )
    terminology_version = c.execute(
        "SELECT value FROM app_meta WHERE key='responsible_person_terminology_version'"
    ).fetchone()
    if not terminology_version or terminology_version[0] != "4":
        c.execute(
            "UPDATE form_fields SET label='Sorumlu Personel', is_system=1 "
            "WHERE field_key='defter_sira_no'"
        )
        c.execute("DELETE FROM form_fields WHERE field_key='personel_adi'")
        for table_name, id_column in (
            ("companies", "sira_no"),
            ("pending_actions", "id"),
        ):
            rows = c.execute(
                f"SELECT {id_column}, data FROM {table_name} WHERE data IS NOT NULL"
            ).fetchall()
            for row_id, raw_data in rows:
                try:
                    record_data = json.loads(raw_data)
                except (TypeError, json.JSONDecodeError):
                    continue
                if not isinstance(record_data, dict):
                    continue
                if "personel_adi" in record_data:
                    record_data.pop("personel_adi")
                    c.execute(
                        f"UPDATE {table_name} SET data=? WHERE {id_column}=?",
                        (json.dumps(record_data, ensure_ascii=False), row_id),
                    )
        c.execute(
            """INSERT INTO app_meta(key,value)
               VALUES('responsible_person_terminology_version','4')
               ON CONFLICT(key) DO UPDATE SET value=excluded.value"""
        )
    instrument_version = c.execute(
        "SELECT value FROM app_meta WHERE key='instrument_items_schema_version'"
    ).fetchone()
    if not instrument_version or instrument_version[0] != "1":
        for table_name, id_column in (
            ("companies", "sira_no"),
            ("pending_actions", "id"),
        ):
            rows = c.execute(
                f"SELECT {id_column}, data FROM {table_name} WHERE data IS NOT NULL"
            ).fetchall()
            for row_id, raw_data in rows:
                try:
                    record_data = json.loads(raw_data)
                except (TypeError, json.JSONDecodeError):
                    continue
                if not isinstance(record_data, dict):
                    continue
                items = normalize_instrument_items(
                    record_data.get("olcu_aletleri"),
                    record_data.get("olcu_aleti_sayisi", ""),
                    record_data.get("olcu_aleti_cinsi", ""),
                )
                if not items:
                    continue
                record_data["olcu_aletleri"] = items
                record_data["olcu_aleti_sayisi"] = instrument_total(items)
                record_data["olcu_aleti_cinsi"] = instrument_summary(items)
                c.execute(
                    f"UPDATE {table_name} SET data=? WHERE {id_column}=?",
                    (json.dumps(record_data, ensure_ascii=False), row_id),
                )
        c.execute(
            """INSERT INTO app_meta(key,value)
               VALUES('instrument_items_schema_version','1')
               ON CONFLICT(key) DO UPDATE SET value=excluded.value"""
        )
    db.commit()
    db.close()

# ---------------------------------------------------------------- yardımcılar
def login_required(f):
    @wraps(f)
    def wrapper(*a, **kw):
        if "user_id" not in session:
            flash("Lütfen önce giriş yapınız.", "warning")
            return redirect(url_for("login"))
        return f(*a, **kw)
    return wrapper

def admin_required(f):
    @wraps(f)
    def wrapper(*a, **kw):
        if session.get("role") != "admin":
            flash("Bu işlem için yönetici (admin) yetkisi gereklidir.", "danger")
            return redirect(url_for("dashboard"))
        return f(*a, **kw)
    return wrapper

def get_fields():
    rows = get_db().execute(
        "SELECT * FROM form_fields ORDER BY position, id").fetchall()
    fields = []
    for r in rows:
        d = dict(r)
        d["options"] = json.loads(r["options"]) if r["options"] else None
        fields.append(d)
    return fields

def input_fields():
    return [f for f in get_fields() if not f["is_system"]]

def amount_value(value):
    if value in (None, ""):
        return 0.0
    text = str(value).strip().replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except (TypeError, ValueError):
        return 0.0

def form_sections():
    fields = input_fields()
    by_key = {field["field_key"]: field for field in fields}
    used = set()
    sections = []
    for title, description, keys in DETAIL_SECTION_KEYS:
        section_fields = [by_key[key] for key in keys if key in by_key]
        if not section_fields:
            continue
        used.update(field["field_key"] for field in section_fields)
        sections.append({
            "title": title,
            "description": description,
            "fields": section_fields,
        })
    remaining = [
        field for field in fields if field["field_key"] not in used
    ]
    if remaining:
        sections.append({
            "title": "Diğer Bilgiler",
            "description": "Forma özel olarak eklenmiş alanlar.",
            "fields": remaining,
        })
    return sections

def parse_tr_date(s):
    s = (s or "").strip()
    for fmt in ("%d.%m.%Y", "%d.%m.%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None

def kalan_sure(company_data):
    d = parse_tr_date(company_data.get("teblig_tarihi", ""))
    if not d:
        return None
    closed_fields = (
        "odeme_tarihi",
        "kesinlesme_tarihi",
        "vergi_bildirim_tarihi",
        "itiraz_vergi_bildirim_tarihi",
        "itiraz_sonucu",
        "vergi_tahsil_tarihi",
    )
    if any(str(company_data.get(key) or "").strip() for key in closed_fields):
        return None
    for key in ("dogrudan_yatirilan", "itiraz_lehine_tutar", "vergi_tahsil_tutari"):
        try:
            if float(str(company_data.get(key) or "0").replace(",", ".")) > 0:
                return None
        except ValueError:
            pass
    source_status = str(company_data.get("kaynak_durumu") or "")
    source_status = source_status.replace("I", "ı").replace("İ", "i").lower()
    unpaid_words = ("yapılmadı", "alınmadı", "alınamadı")
    paid_words = ("yapıldı", "alındı", "yapılmış")
    if "iptal" in source_status or (
        any(word in source_status for word in paid_words)
        and not any(word in source_status for word in unpaid_words)
    ):
        return None
    return ((d + timedelta(days=30)) - date.today()).days

def sync_instrument_data(data):
    """Keep paired instrument items and Excel-compatible F/G values aligned."""
    items = normalize_instrument_items(
        data.get("olcu_aletleri"),
        data.get("olcu_aleti_sayisi", ""),
        data.get("olcu_aleti_cinsi", ""),
    )
    data["olcu_aletleri"] = items
    data["olcu_aleti_sayisi"] = instrument_total(items)
    data["olcu_aleti_cinsi"] = instrument_summary(items)
    return data

def instrument_items_for_form(data):
    raw_items = data.get("olcu_aletleri") if hasattr(data, "get") else None
    if isinstance(raw_items, list) and raw_items:
        return [
            {
                "cinsi": str(item.get("cinsi", "")) if isinstance(item, dict) else "",
                "sayisi": str(item.get("sayisi", "")) if isinstance(item, dict) else "",
            }
            for item in raw_items
        ]
    items = normalize_instrument_items(
        quantity=data.get("olcu_aleti_sayisi", "") if hasattr(data, "get") else "",
        kind=data.get("olcu_aleti_cinsi", "") if hasattr(data, "get") else "",
    )
    return items or [{"cinsi": "", "sayisi": ""}]

def instrument_type_options():
    field = next(
        (
            field for field in get_fields()
            if field["field_key"] == "olcu_aleti_cinsi"
        ),
        None,
    )
    return field["options"] if field and field["options"] else []

def form_data_for_render(form):
    data = form.to_dict()
    kinds = form.getlist("olcu_aleti_cinsi")
    quantities = form.getlist("olcu_aleti_sayisi")
    data["olcu_aletleri"] = [
        {
            "cinsi": kinds[index] if index < len(kinds) else "",
            "sayisi": quantities[index] if index < len(quantities) else "",
        }
        for index in range(max(len(kinds), len(quantities), 1))
    ]
    return data

def validate_field(field, raw):
    """Alan doğrulama. (deger, TÜRKÇE hata mesajı) döner."""
    val = (raw or "").strip()
    if not val:
        if field["required"]:
            return None, "Bu alanın doldurulması zorunludur."
        return "", None
    dt = field["datatype"]
    key = field["field_key"]
    if key == "tablet_tutanak_no":
        if not re.fullmatch(r"\d{2}-[A-Za-z0-9]{8}", val):
            return None, ("Biçim hatalı: “41-E7E8EC31” örneğindeki gibi olmalıdır "
                          "(2 rakam, tire, ardından 8 harf/rakam).")
        return val.upper(), None
    if key == "ebys":
        if not re.fullmatch(r"\d{2}\.\d{2}\.\d{4}-\d{7}", val) or not parse_tr_date(val.split("-")[0]):
            return None, ("Biçim hatalı: “30.12.2024-5599177” örneğindeki gibi olmalıdır "
                          "(geçerli tarih, tire, ardından 7 rakam).")
        return val, None
    if dt.startswith("TARİH") or dt.startswith("DATETIME"):
        if not parse_tr_date(val):
            return None, "Geçerli bir tarih giriniz. Biçim: GG.AA.YYYY (örn. 4.10.2026)."
        return val, None
    if dt == "INT":
        if not re.fullmatch(r"-?\d+", val):
            return None, "Bu alana yalnızca tam sayı girilebilir (örn. 3)."
        if key == "yil" and not (1900 <= int(val) <= 2100):
            return None, "Yıl 1900 ile 2100 arasında olmalıdır (örn. 2020)."
        return val, None
    if dt in PARA_TIPLERI:
        # Para alanları: rakamlar arasında boşluğa izin verilir ("2 345 234")
        temiz = val.replace(" ", "")
        if not re.fullmatch(r"-?\d+([.,]\d+)?", temiz):
            return None, ("Geçerli bir tutar giriniz (örn. 16000 veya 2 345 234). "
                          "Harf veya özel karakter kullanılamaz.")
        return temiz.replace(",", "."), None
    if dt == "BOOLEAN":
        if val not in ("Evet", "Hayır"):
            return None, "Lütfen Evet veya Hayır seçiniz."
        return val, None
    if dt == "AÇILIR LİSTE" and field["options"] and val not in field["options"]:
        return None, "Lütfen listeden geçerli bir seçenek seçiniz."
    return val, None

def collect_form(form):
    """Şirket formunu doğrular. (veri, {alan: türkçe hata}) döner."""
    data, errors = {}, {}
    for f in input_fields():
        if f["field_key"] in ("olcu_aleti_sayisi", "olcu_aleti_cinsi"):
            continue
        val, err = validate_field(f, form.get(f["field_key"]))
        if err:
            errors[f["field_key"]] = err
        else:
            data[f["field_key"]] = val

    kinds = form.getlist("olcu_aleti_cinsi")
    quantities = form.getlist("olcu_aleti_sayisi")
    option_values = set(instrument_type_options())
    instrument_items = []
    instrument_errors = []
    instrument_row_count = max(len(kinds), len(quantities))
    if instrument_row_count > 50:
        instrument_errors.append("Bir kayda en fazla 50 ölçü aleti kalemi eklenebilir.")
    for index in range(min(instrument_row_count, 50)):
        kind = kinds[index].strip() if index < len(kinds) else ""
        quantity = quantities[index].strip() if index < len(quantities) else ""
        if not kind and not quantity:
            continue
        row_label = f"{index + 1}. ölçü aleti"
        if not kind:
            instrument_errors.append(f"{row_label}: cins seçiniz.")
            continue
        if option_values and kind not in option_values:
            instrument_errors.append(f"{row_label}: listeden geçerli bir cins seçiniz.")
            continue
        if not re.fullmatch(r"\d+", quantity or "") or int(quantity) <= 0:
            instrument_errors.append(f"{row_label}: sayı 1 veya daha büyük olmalıdır.")
            continue
        instrument_items.append({"cinsi": kind, "sayisi": int(quantity)})

    instrument_items = normalize_instrument_items(instrument_items)
    if not instrument_items and not instrument_errors:
        instrument_errors.append("En az bir ölçü aleti cinsi ve sayısı giriniz.")
    if instrument_errors:
        errors["olcu_aletleri"] = " ".join(instrument_errors)
    else:
        data["olcu_aletleri"] = instrument_items
        sync_instrument_data(data)
    return data, errors

def notify(user_id, message, *, title=None, kind="general", priority="info",
           action_url=None, dedupe_key=None, metadata=None):
    get_db().execute(
        """INSERT INTO notifications
           (user_id,title,message,kind,priority,action_url,dedupe_key,metadata)
           VALUES(?,?,?,?,?,?,?,?)""",
        (
            user_id, title, message, kind, priority, action_url, dedupe_key,
            json.dumps(metadata, ensure_ascii=False) if metadata else None,
        ),
    )

def notify_admins(message, exclude=None):
    db = get_db()
    for r in db.execute("SELECT id FROM users WHERE role='admin'"):
        if r["id"] != exclude:
            notify(r["id"], message)

def _sync_grouped_notification(user_id, *, dedupe_key, count, title, message,
                               kind, priority, action_url, mark_unread=True):
    db = get_db()
    current = db.execute(
        """SELECT id,is_read,metadata FROM notifications
           WHERE user_id=? AND dedupe_key=? AND resolved_at IS NULL""",
        (user_id, dedupe_key),
    ).fetchone()
    if count <= 0:
        if current:
            db.execute(
                """UPDATE notifications
                   SET resolved_at=datetime('now','localtime'), is_read=1
                   WHERE id=?""",
                (current["id"],),
            )
        return

    old_count = 0
    if current and current["metadata"]:
        try:
            old_count = int(json.loads(current["metadata"]).get("count") or 0)
        except (TypeError, ValueError, json.JSONDecodeError):
            old_count = 0
    should_mark_unread = mark_unread and (not current or count > old_count)
    metadata = json.dumps({"count": count}, ensure_ascii=False)
    if current:
        db.execute(
            """UPDATE notifications
               SET title=?,message=?,kind=?,priority=?,action_url=?,metadata=?,
                   is_read=CASE WHEN ? THEN 0 ELSE is_read END
               WHERE id=?""",
            (
                title, message, kind, priority, action_url, metadata,
                int(should_mark_unread), current["id"],
            ),
        )
    else:
        notify(
            user_id, message, title=title, kind=kind, priority=priority,
            action_url=action_url, dedupe_key=dedupe_key,
            metadata={"count": count},
        )

def sync_admin_approval_notifications(mark_unread=True):
    db = get_db()
    count = db.execute(
        "SELECT COUNT(*) c FROM pending_actions WHERE status='bekliyor'"
    ).fetchone()["c"]
    for admin in db.execute("SELECT id FROM users WHERE role='admin'"):
        _sync_grouped_notification(
            admin["id"],
            dedupe_key="pending-approvals",
            count=count,
            title="Onay bekleyen işlemler",
            message=f"{count} kayıt işlemi incelemenizi bekliyor.",
            kind="approval",
            priority="action",
            action_url="/onaylar",
            mark_unread=mark_unread,
        )

def sync_deadline_notifications(user_id):
    if session.get("role") != "admin":
        return
    companies = _load_companies()
    expired = sum(
        company["status"]["key"] == "sure-doldu" for company in companies
    )
    upcoming = sum(
        company["status"]["key"] == "sure-isliyor"
        and company["kalan"] is not None
        and 0 <= company["kalan"] <= 7
        for company in companies
    )
    _sync_grouped_notification(
        user_id,
        dedupe_key="deadline-expired",
        count=expired,
        title="Süresi dolan kayıtlar",
        message=f"{expired} kayıt için süre dolmuş görünüyor.",
        kind="deadline",
        priority="critical",
        action_url="/sirketler?durum=sure-doldu",
    )
    _sync_grouped_notification(
        user_id,
        dedupe_key="deadline-upcoming",
        count=upcoming,
        title="Süresi yaklaşan kayıtlar",
        message=f"{upcoming} kaydın süresi 7 gün içinde dolacak.",
        kind="deadline",
        priority="action",
        action_url="/sirketler?durum=sure-isliyor",
    )

def _clean_attachment_name(filename):
    """Keep the user's display name, but never allow it to become a disk path."""
    name = os.path.basename((filename or "").replace("\\", "/")).strip()
    name = "".join(ch for ch in name if ch >= " " and ch != "\x7f")
    return name[:255]

def _prepare_attachments(files):
    """Validate uploaded files without writing them to disk."""
    uploads = [f for f in files if f and (f.filename or "").strip()]
    if len(uploads) > MAX_ATTACHMENT_COUNT:
        return [], f"Bir işlemde en fazla {MAX_ATTACHMENT_COUNT} dosya yükleyebilirsiniz."

    prepared = []
    for upload in uploads:
        original_name = _clean_attachment_name(upload.filename)
        extension = os.path.splitext(original_name)[1].lower()
        if not original_name or extension not in ALLOWED_ATTACHMENT_EXTENSIONS:
            shown_name = original_name or "adsız dosya"
            return [], (
                f"“{shown_name}” dosya türüne izin verilmiyor. "
                "Belge, tablo, sunu, arşiv, fotoğraf veya video yükleyiniz."
            )
        try:
            upload.stream.seek(0, os.SEEK_END)
            file_size = upload.stream.tell()
            upload.stream.seek(0)
        except (AttributeError, OSError):
            return [], f"“{original_name}” dosyasının boyutu okunamadı."
        if file_size <= 0:
            return [], f"“{original_name}” boş bir dosya olduğu için yüklenemedi."
        if file_size > MAX_ATTACHMENT_SIZE:
            return [], f"“{original_name}” 100 MB dosya sınırını aşıyor."
        guessed_type = mimetypes.guess_type(original_name)[0]
        prepared.append({
            "file": upload,
            "original_name": original_name,
            "stored_name": f"{uuid.uuid4().hex}{extension}",
            "mime_type": guessed_type or "application/octet-stream",
            "file_size": file_size,
        })
    return prepared, None

def _delete_stored_files(stored_names):
    """Delete only UUID-named files that were resolved from database rows."""
    for stored_name in stored_names:
        if not re.fullmatch(r"[0-9a-f]{32}\.[a-z0-9]+", stored_name or ""):
            app.logger.warning("Güvenli olmayan ek dosya adı atlandı: %r", stored_name)
            continue
        path = os.path.join(UPLOAD_DIR, stored_name)
        try:
            os.remove(path)
        except FileNotFoundError:
            pass
        except OSError:
            app.logger.exception("Ek dosya diskten silinemedi: %s", stored_name)

def _save_prepared_attachments(prepared, company_sira_no=None, pending_action_id=None):
    if not prepared:
        return []
    if company_sira_no is None and pending_action_id is None:
        raise ValueError("Ek dosya için kayıt veya onay talebi belirtilmelidir.")

    os.makedirs(UPLOAD_DIR, mode=0o700, exist_ok=True)
    saved_names = []
    try:
        for item in prepared:
            path = os.path.join(UPLOAD_DIR, item["stored_name"])
            item["file"].save(path)
            os.chmod(path, 0o600)
            saved_names.append(item["stored_name"])
            get_db().execute(
                """INSERT INTO attachments
                   (company_sira_no,pending_action_id,original_name,stored_name,
                    mime_type,file_size,uploaded_by)
                   VALUES(?,?,?,?,?,?,?)""",
                (
                    company_sira_no,
                    pending_action_id,
                    item["original_name"],
                    item["stored_name"],
                    item["mime_type"],
                    item["file_size"],
                    session.get("user_id"),
                ),
            )
    except Exception:
        _delete_stored_files(saved_names)
        raise
    return saved_names

def _attachment_rows(where_column, value):
    if where_column not in ("company_sira_no", "pending_action_id"):
        raise ValueError("Geçersiz ek dosya sorgusu.")
    rows = [
        dict(row) for row in get_db().execute(
            f"SELECT * FROM attachments WHERE {where_column}=? ORDER BY id",
            (value,),
        ).fetchall()
    ]
    for row in rows:
        row["can_preview"] = (
            os.path.splitext(row["stored_name"])[1].lower()
            in INLINE_ATTACHMENT_EXTENSIONS
        )
    return rows

def _attachments_for_company(sira_no):
    return _attachment_rows("company_sira_no", sira_no)

def _attachments_for_pending(pending_id):
    return _attachment_rows("pending_action_id", pending_id)

@app.template_filter("filesize")
def format_file_size(size):
    value = float(size or 0)
    for unit in ("B", "KB", "MB", "GB"):
        if value < 1024 or unit == "GB":
            return f"{value:.0f} {unit}" if unit == "B" else f"{value:.1f} {unit}"
        value /= 1024
    return f"{value:.1f} GB"

def apply_company_action(action_type, sira_no, data_json, actor_name):
    db = get_db()
    if action_type == "ekle":
        d = json.loads(data_json)
        d["defter_sira_no"] = d.get("defter_sira_no") or actor_name
        cursor = db.execute(
            "INSERT INTO companies(data,created_by) VALUES(?,?)",
            (json.dumps(d, ensure_ascii=False), actor_name),
        )
        return {"sira_no": cursor.lastrowid, "deleted_files": []}
    elif action_type == "duzenle":
        db.execute("UPDATE companies SET data=?, updated_at=datetime('now','localtime') WHERE sira_no=?",
                   (data_json, sira_no))
        return {"sira_no": sira_no, "deleted_files": []}
    elif action_type == "sil":
        stored_names = [
            row["stored_name"] for row in db.execute(
                "SELECT stored_name FROM attachments WHERE company_sira_no=?",
                (sira_no,),
            ).fetchall()
        ]
        db.execute("DELETE FROM attachments WHERE company_sira_no=?", (sira_no,))
        db.execute("DELETE FROM companies WHERE sira_no=?", (sira_no,))
        return {"sira_no": sira_no, "deleted_files": stored_names}
    raise ValueError(f"Geçersiz şirket işlemi: {action_type}")

@app.context_processor
def inject_globals():
    unread = pending = 0
    if "user_id" in session:
        db = get_db()
        unread = db.execute(
            """SELECT COUNT(*) c FROM notifications
               WHERE user_id=? AND is_read=0 AND resolved_at IS NULL""",
            (session["user_id"],)).fetchone()["c"]
        if session.get("role") == "admin":
            pending = db.execute(
                "SELECT COUNT(*) c FROM pending_actions WHERE status='bekliyor'").fetchone()["c"]
    return dict(
        unread_count=unread,
        pending_count=pending,
        instrument_items_for_form=instrument_items_for_form,
        instrument_type_options=instrument_type_options,
    )

# ---------------------------------------------------------------- kimlik
@app.route("/")
def index():
    return redirect(url_for("dashboard" if "user_id" in session else "login"))

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        ident = request.form.get("ident", "").strip()
        pwd = request.form.get("password", "")
        u = get_db().execute(
            "SELECT * FROM users WHERE email=? OR name=?", (ident, ident)).fetchone()
        if u and check_password_hash(u["password_hash"], pwd):
            session.clear()
            session["user_id"], session["user_name"], session["role"] = u["id"], u["name"], u["role"]
            flash(f"Hoş geldiniz, {u['name']}!", "success")
            return redirect(url_for("dashboard"))
        flash("Kullanıcı adı/e-posta veya şifre hatalı.", "danger")
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    flash("Başarıyla çıkış yaptınız.", "success")
    return redirect(url_for("login"))

@app.route("/sifremi-unuttum", methods=["GET", "POST"])
def forgot():
    if request.method == "POST":
        email = request.form.get("email", "").strip()
        u = get_db().execute("SELECT * FROM users WHERE email=?", (email,)).fetchone()
        if u:
            token = serializer.dumps(email, salt="sifre-sifirla")
            link = url_for("reset", token=token, _external=True)
            print(f"[ŞİFRE SIFIRLAMA] {email} -> {link}")
            flash("Şifre sıfırlama bağlantısı e-posta adresinize gönderildi.", "success")
            flash(f"(Yerel mod) Sıfırlama bağlantınız: {link}", "info")
        else:
            flash("Bu e-posta adresi ile kayıtlı kullanıcı bulunamadı.", "danger")
    return render_template("forgot.html")

@app.route("/sifre-sifirla/<token>", methods=["GET", "POST"])
def reset(token):
    try:
        email = serializer.loads(token, salt="sifre-sifirla", max_age=3600)
    except (BadSignature, SignatureExpired):
        flash("Bağlantı geçersiz veya süresi dolmuş.", "danger")
        return redirect(url_for("forgot"))
    if request.method == "POST":
        p1, p2 = request.form.get("password", ""), request.form.get("password2", "")
        if len(p1) < 6:
            flash("Şifre en az 6 karakter olmalıdır.", "danger")
        elif p1 != p2:
            flash("Girdiğiniz şifreler birbiriyle eşleşmiyor.", "danger")
        else:
            db = get_db()
            db.execute("UPDATE users SET password_hash=? WHERE email=?",
                       (generate_password_hash(p1), email))
            db.commit()
            flash("Şifreniz güncellendi, yeni şifrenizle giriş yapabilirsiniz.", "success")
            return redirect(url_for("login"))
    return render_template("reset.html", token=token)

# ---------------------------------------------------------------- pano
@app.route("/panel")
@login_required
def dashboard():
    companies = _load_companies()
    current_year = str(date.today().year)
    total_amount = 0.0
    collected_amount = 0.0
    year_counts, year_amounts, legal_basis_counts = {}, {}, {}

    for company in companies:
        data = company["data"]
        penalty_amount = amount_value(data.get("ceza_tutari"))
        total_amount += penalty_amount
        collected_amount += (
            amount_value(data.get("dogrudan_yatirilan"))
            + amount_value(data.get("vergi_tahsil_tutari"))
        )

        year = str(data.get("yil") or "Belirsiz")
        year_counts[year] = year_counts.get(year, 0) + 1
        year_amounts[year] = year_amounts.get(year, 0.0) + penalty_amount

        legal_basis = str(data.get("hukuki_dayanak") or "Belirsiz")
        legal_basis_counts[legal_basis] = (
            legal_basis_counts.get(legal_basis, 0) + 1
        )

    attention_records = [
        company for company in companies
        if company["kalan"] is not None and company["kalan"] <= 3
    ]
    attention_records.sort(key=lambda company: company["kalan"])

    status_groups = [
        {
            "key": "sure-isliyor",
            "label": "Süresi işleyen",
            "count": sum(
                company["status"]["key"] == "sure-isliyor"
                for company in companies
            ),
            "tone": "warning",
        },
        {
            "key": "sure-doldu",
            "label": "Süresi dolan",
            "count": sum(
                company["status"]["key"] == "sure-doldu"
                for company in companies
            ),
            "tone": "danger",
        },
        {
            "key": "odendi",
            "label": "Ödenen",
            "count": sum(
                company["status"]["key"] == "odendi"
                for company in companies
            ),
            "tone": "success",
        },
    ]
    grouped_status_total = sum(group["count"] for group in status_groups)
    status_groups.append({
        "key": "",
        "label": "Diğer süreçler",
        "count": max(len(companies) - grouped_status_total, 0),
        "tone": "neutral",
    })
    for group in status_groups:
        group["percent"] = (
            round((group["count"] / len(companies)) * 100)
            if companies else 0
        )

    sorted_years = sorted(
        year_counts,
        key=lambda year: (
            year == "Belirsiz",
            int(year) if year.isdigit() else 9999,
            tr_kucuk(year),
        ),
    )
    top_legal_bases = sorted(
        legal_basis_counts.items(), key=lambda item: (-item[1], tr_kucuk(item[0]))
    )[:6]
    chart_data = {
        "years": sorted_years,
        "year_counts": [year_counts[year] for year in sorted_years],
        "year_amounts": [year_amounts[year] for year in sorted_years],
        "legal_basis_labels": [item[0] for item in top_legal_bases],
        "legal_basis_counts": [item[1] for item in top_legal_bases],
        "status_labels": [group["label"] for group in status_groups],
        "status_counts": [group["count"] for group in status_groups],
    }

    return render_template(
        "dashboard.html",
        toplam=len(companies),
        toplam_ceza=total_amount,
        tahsil_edilen=collected_amount,
        bu_yil=year_counts.get(current_year, 0),
        uyari_sayisi=len(attention_records),
        status_groups=status_groups,
        chart_data=chart_data,
    )

# ---------------------------------------------------------------- kullanıcı yönetimi (admin)
@app.route("/personel", methods=["GET", "POST"])
@login_required
@admin_required
def personel():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        email = request.form.get("email", "").strip()
        pwd = request.form.get("password", "")
        role = request.form.get("role", "personel")
        if not name or not email or len(pwd) < 6:
            flash("Ad ve e-posta zorunludur; şifre en az 6 karakter olmalıdır.", "danger")
        else:
            db = get_db()
            try:
                db.execute("INSERT INTO users(name,email,password_hash,role) VALUES(?,?,?,?)",
                           (name, email, generate_password_hash(pwd), role))
                db.commit()
                flash(f"“{name}” kullanıcısı başarıyla eklendi.", "success")
                return redirect(url_for("personel"))
            except sqlite3.IntegrityError:
                flash("Bu e-posta adresi zaten kayıtlı.", "danger")
    users = get_db().execute("SELECT * FROM users ORDER BY id").fetchall()
    return render_template("personel.html", users=users)

@app.route("/personel/<int:uid>/duzenle", methods=["POST"])
@login_required
@admin_required
def personel_duzenle(uid):
    db = get_db()
    u = db.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
    if not u:
        abort(404)
    name = request.form.get("name", "").strip() or u["name"]
    email = request.form.get("email", "").strip() or u["email"]
    role = request.form.get("role", u["role"])
    pwd = request.form.get("password", "")
    if uid == session["user_id"] and role != "admin":
        flash("Kendi yönetici yetkinizi kaldıramazsınız.", "danger")
        return redirect(url_for("personel"))
    try:
        db.execute("UPDATE users SET name=?, email=?, role=? WHERE id=?",
                   (name, email, role, uid))
        if pwd:
            if len(pwd) < 6:
                flash("Yeni şifre en az 6 karakter olmalıdır; şifre değiştirilmedi.", "warning")
            else:
                db.execute("UPDATE users SET password_hash=? WHERE id=?",
                           (generate_password_hash(pwd), uid))
        db.commit()
        if uid == session["user_id"]:
            session["user_name"] = name
        notify(uid, f"Hesap bilgileriniz yönetici ({session['user_name']}) tarafından güncellendi.")
        db.commit()
        flash(f"“{name}” kullanıcısının bilgileri güncellendi.", "success")
    except sqlite3.IntegrityError:
        flash("Bu e-posta adresi başka bir kullanıcıya ait.", "danger")
    return redirect(url_for("personel"))

@app.route("/personel/<int:uid>/sil", methods=["POST"])
@login_required
@admin_required
def personel_sil(uid):
    db = get_db()
    u = db.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
    if not u:
        abort(404)
    if uid == session["user_id"]:
        flash("Kendi hesabınızı silemezsiniz.", "danger")
        return redirect(url_for("personel"))
    if u["role"] == "admin":
        kalan_admin = db.execute(
            "SELECT COUNT(*) c FROM users WHERE role='admin'").fetchone()["c"]
        if kalan_admin <= 1:
            flash("Sistemdeki son yönetici silinemez.", "danger")
            return redirect(url_for("personel"))
    db.execute("DELETE FROM users WHERE id=?", (uid,))
    db.commit()
    flash(f"“{u['name']}” kullanıcısı silindi.", "warning")
    return redirect(url_for("personel"))

# ---------------------------------------------------------------- profil (herkes)
@app.route("/profil", methods=["GET", "POST"])
@login_required
def profil():
    db = get_db()
    u = db.execute("SELECT * FROM users WHERE id=?", (session["user_id"],)).fetchone()
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        email = request.form.get("email", "").strip()
        pwd = request.form.get("password", "")
        pwd2 = request.form.get("password2", "")
        if not name or not email:
            flash("Ad ve e-posta alanları boş bırakılamaz.", "danger")
        elif pwd and len(pwd) < 6:
            flash("Yeni şifre en az 6 karakter olmalıdır.", "danger")
        elif pwd and pwd != pwd2:
            flash("Girdiğiniz yeni şifreler birbiriyle eşleşmiyor.", "danger")
        else:
            degisen = []
            if name != u["name"]:
                degisen.append("ad")
            if email != u["email"]:
                degisen.append("e-posta")
            if pwd:
                degisen.append("şifre")
            try:
                db.execute("UPDATE users SET name=?, email=? WHERE id=?",
                           (name, email, session["user_id"]))
                if pwd:
                    db.execute("UPDATE users SET password_hash=? WHERE id=?",
                               (generate_password_hash(pwd), session["user_id"]))
                if degisen:
                    notify_admins(f"{u['name']} kendi profil bilgilerini güncelledi "
                                  f"({', '.join(degisen)}).", exclude=session["user_id"])
                db.commit()
                session["user_name"] = name
                flash("Profil bilgileriniz başarıyla güncellendi.", "success")
                return redirect(url_for("profil"))
            except sqlite3.IntegrityError:
                flash("Bu e-posta adresi başka bir kullanıcıya ait.", "danger")
        u = db.execute("SELECT * FROM users WHERE id=?", (session["user_id"],)).fetchone()
    return render_template("profil.html", user=u)

# ---------------------------------------------------------------- form düzenle (admin)
@app.route("/form-duzenle")
@login_required
@admin_required
def form_duzenle():
    fields = get_fields()
    return render_template(
        "form_duzenle.html",
        fields=fields,
        official_fields=fields,
        excel_columns=REAL_EXCEL_COLUMNS,
        datatypes=DATATYPES,
    )

@app.route("/form-duzenle/ekle", methods=["POST"])
@login_required
@admin_required
def form_alan_ekle():
    label = request.form.get("label", "").strip()
    dt = request.form.get("datatype", "TEXT")
    req = 1 if request.form.get("required") == "on" else 0
    opts = [o.strip() for o in request.form.get("options", "").splitlines() if o.strip()]
    if not label:
        flash("Sütun adı boş olamaz.", "danger")
        return redirect(url_for("form_duzenle"))
    key = re.sub(r"[^a-z0-9_]", "_", label.lower()
                 .translate(str.maketrans("çğıöşü", "cgiosu"))).strip("_") or "alan"
    db = get_db()
    base, i = key, 1
    while db.execute("SELECT 1 FROM form_fields WHERE field_key=?", (key,)).fetchone():
        i += 1
        key = f"{base}_{i}"
    pos = db.execute("SELECT COALESCE(MAX(position),0)+1 p FROM form_fields").fetchone()["p"]
    db.execute("""INSERT INTO form_fields(field_key,label,datatype,required,options,position)
                  VALUES(?,?,?,?,?,?)""",
               (key, label, dt, req,
                json.dumps(opts, ensure_ascii=False) if opts else None, pos))
    db.commit()
    flash(f"“{label}” sütunu forma eklendi.", "success")
    return redirect(url_for("form_duzenle"))

@app.route("/form-duzenle/<int:fid>/guncelle", methods=["POST"])
@login_required
@admin_required
def form_alan_guncelle(fid):
    db = get_db()
    f = db.execute("SELECT * FROM form_fields WHERE id=?", (fid,)).fetchone()
    if not f:
        abort(404)
    if f["is_system"]:
        flash("Sistem alanları (Sıra No, Personel Adı, Kalan Süre) düzenlenemez.", "danger")
        return redirect(url_for("form_duzenle"))
    label = request.form.get("label", "").strip() or f["label"]
    dt = request.form.get("datatype", f["datatype"])
    req = 1 if request.form.get("required") == "on" else 0
    opts = [o.strip() for o in request.form.get("options", "").splitlines() if o.strip()]
    db.execute("UPDATE form_fields SET label=?, datatype=?, required=?, options=? WHERE id=?",
               (label, dt, req, json.dumps(opts, ensure_ascii=False) if opts else None, fid))
    db.commit()
    flash(f"“{label}” sütunu güncellendi.", "success")
    return redirect(url_for("form_duzenle"))

@app.route("/form-duzenle/<int:fid>/sil", methods=["POST"])
@login_required
@admin_required
def form_alan_sil(fid):
    db = get_db()
    f = db.execute("SELECT * FROM form_fields WHERE id=?", (fid,)).fetchone()
    if not f:
        abort(404)
    if f["is_system"]:
        flash("Sistem alanları silinemez.", "danger")
    elif f["field_key"] in REAL_EXCEL_COLUMNS:
        flash(
            f"“{f['label']}” gerçek Excel dosyasının {REAL_EXCEL_COLUMNS[f['field_key']]} "
            "sütununa eşlenmiştir ve silinemez. Alanı düzenleyebilirsiniz.",
            "danger",
        )
    else:
        db.execute("DELETE FROM form_fields WHERE id=?", (fid,))
        db.commit()
        flash(f"“{f['label']}” sütunu formdan silindi.", "warning")
    return redirect(url_for("form_duzenle"))

# ---------------------------------------------------------------- gerçek Excel verisi aktarımı (admin)
def _cleanup_import_files():
    os.makedirs(IMPORT_TMP_DIR, mode=0o700, exist_ok=True)
    cutoff = time.time() - (2 * 60 * 60)
    for name in os.listdir(IMPORT_TMP_DIR):
        if not re.fullmatch(r"[0-9a-f]{32}\.xlsx", name):
            continue
        path = os.path.join(IMPORT_TMP_DIR, name)
        try:
            if os.path.getmtime(path) < cutoff:
                os.unlink(path)
        except OSError:
            pass


def _import_path(token):
    if not re.fullmatch(r"[0-9a-f]{32}", token or ""):
        return None
    return os.path.join(IMPORT_TMP_DIR, f"{token}.xlsx")


def _remove_import_file(token):
    path = _import_path(token)
    if path and os.path.isfile(path):
        try:
            os.unlink(path)
        except OSError:
            pass


def _ensure_import_fields(db):
    position = db.execute(
        "SELECT COALESCE(MAX(position), -1) + 1 p FROM form_fields"
    ).fetchone()["p"]
    for key, label, datatype, required, options, is_system in IMPORT_FIELD_DEFINITIONS:
        if db.execute(
            "SELECT 1 FROM form_fields WHERE field_key=?", (key,)
        ).fetchone():
            continue
        db.execute(
            """INSERT INTO form_fields
               (field_key,label,datatype,required,options,is_system,position)
               VALUES(?,?,?,?,?,?,?)""",
            (
                key,
                label,
                datatype,
                required,
                json.dumps(options, ensure_ascii=False) if options else None,
                is_system,
                position,
            ),
        )
        position += 1


def _merge_import_options(db, option_values):
    for key, imported_values in option_values.items():
        field = db.execute(
            "SELECT options FROM form_fields WHERE field_key=?", (key,)
        ).fetchone()
        if not field:
            continue
        existing = json.loads(field["options"]) if field["options"] else []
        seen = set(existing)
        merged = list(existing)
        for value in imported_values:
            if value not in seen:
                merged.append(value)
                seen.add(value)
        db.execute(
            "UPDATE form_fields SET options=? WHERE field_key=?",
            (json.dumps(merged, ensure_ascii=False), key),
        )


def _create_import_backup(db):
    os.makedirs(BACKUP_DIR, mode=0o700, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_path = os.path.join(BACKUP_DIR, f"ipc-before-import-{stamp}.db")
    db.commit()
    with sqlite3.connect(backup_path) as backup:
        db.backup(backup)
    return backup_path


def _prepare_preview(parsed):
    db = get_db()
    existing_ids = {
        row["sira_no"] for row in db.execute("SELECT sira_no FROM companies")
    }
    source_ids = {record["sira_no"] for record in parsed["records"]}
    previous = db.execute(
        """SELECT filename, imported_at, inserted_rows, updated_rows
           FROM import_history WHERE file_sha256=?
           ORDER BY id DESC LIMIT 1""",
        (parsed["digest"],),
    ).fetchone()
    preview = dict(parsed)
    preview.pop("records", None)
    preview.pop("option_values", None)
    preview["existing_total"] = len(existing_ids)
    preview["duplicate_count"] = len(source_ids & existing_ids)
    preview["new_count"] = len(source_ids - existing_ids)
    preview["previous_import"] = dict(previous) if previous else None
    return preview


@app.route("/veri-aktar", methods=["GET", "POST"])
@login_required
@admin_required
def veri_aktar():
    _cleanup_import_files()
    if request.method == "POST":
        upload = request.files.get("workbook")
        original_name = os.path.basename((upload.filename if upload else "") or "")
        if not upload or not original_name:
            flash("Lütfen aktarılacak Excel dosyasını seçiniz.", "danger")
            return render_template("veri_aktar.html", preview=None, result=None)
        if os.path.splitext(original_name)[1].lower() not in (".xlsx", ".xlsm"):
            flash("Yalnızca .xlsx veya .xlsm dosyaları aktarılabilir.", "danger")
            return render_template("veri_aktar.html", preview=None, result=None)

        token = uuid.uuid4().hex
        path = _import_path(token)
        os.makedirs(IMPORT_TMP_DIR, mode=0o700, exist_ok=True)
        upload.save(path)
        try:
            parsed = parse_workbook(path)
        except ImportWorkbookError as exc:
            _remove_import_file(token)
            flash(str(exc), "danger")
            return render_template("veri_aktar.html", preview=None, result=None)

        parsed["filename"] = original_name
        session["ipc_import"] = {
            "token": token,
            "digest": parsed["digest"],
            "filename": original_name,
        }
        return render_template(
            "veri_aktar.html",
            preview=_prepare_preview(parsed),
            result=None,
            token=token,
        )
    return render_template("veri_aktar.html", preview=None, result=None)


@app.route("/veri-aktar/onayla", methods=["POST"])
@login_required
@admin_required
def veri_aktar_onayla():
    state = session.get("ipc_import") or {}
    token = request.form.get("token", "")
    conflict_mode = request.form.get("conflict_mode", "skip")
    if conflict_mode not in ("skip", "update"):
        abort(400)
    if token != state.get("token"):
        flash("Aktarım önizlemesinin süresi dolmuş. Dosyayı yeniden seçiniz.", "danger")
        return redirect(url_for("veri_aktar"))
    path = _import_path(token)
    if not path or not os.path.isfile(path):
        session.pop("ipc_import", None)
        flash("Geçici aktarım dosyası bulunamadı. Dosyayı yeniden seçiniz.", "danger")
        return redirect(url_for("veri_aktar"))

    try:
        parsed = parse_workbook(path)
    except ImportWorkbookError as exc:
        _remove_import_file(token)
        session.pop("ipc_import", None)
        flash(str(exc), "danger")
        return redirect(url_for("veri_aktar"))
    if parsed["digest"] != state.get("digest") or file_sha256(path) != state.get("digest"):
        _remove_import_file(token)
        session.pop("ipc_import", None)
        flash("Önizlenen dosya değiştiği için aktarım iptal edildi.", "danger")
        return redirect(url_for("veri_aktar"))

    db = get_db()
    try:
        backup_path = _create_import_backup(db)
    except (OSError, sqlite3.Error) as exc:
        flash(f"Veritabanı yedeği oluşturulamadığı için aktarım başlatılmadı: {exc}", "danger")
        return redirect(url_for("veri_aktar"))

    inserted = updated = skipped = 0
    try:
        db.execute("BEGIN IMMEDIATE")
        _ensure_import_fields(db)
        _merge_import_options(db, parsed["option_values"])
        for record in parsed["records"]:
            sira_no = record["sira_no"]
            existing = db.execute(
                "SELECT data FROM companies WHERE sira_no=?", (sira_no,)
            ).fetchone()
            incoming = dict(record["data"])
            if existing:
                if conflict_mode == "skip":
                    skipped += 1
                    continue
                current = json.loads(existing["data"])
                current.update(incoming)
                db.execute(
                    """UPDATE companies
                       SET data=?, updated_at=datetime('now','localtime'),
                           source_file_hash=?, source_row=?
                       WHERE sira_no=?""",
                    (
                        json.dumps(current, ensure_ascii=False),
                        parsed["digest"],
                        record["source_row"],
                        sira_no,
                    ),
                )
                updated += 1
            else:
                db.execute(
                    """INSERT INTO companies
                       (sira_no,data,created_by,source_file_hash,source_row)
                       VALUES(?,?,?,?,?)""",
                    (
                        sira_no,
                        json.dumps(incoming, ensure_ascii=False),
                        "Excel veri aktarımı",
                        parsed["digest"],
                        record["source_row"],
                    ),
                )
                inserted += 1

        db.execute(
            """INSERT INTO import_history
               (filename,file_sha256,sheet_name,source_rows,inserted_rows,
                updated_rows,skipped_rows,warning_count,conflict_mode,
                backup_path,imported_by)
               VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
            (
                state.get("filename") or parsed["filename"],
                parsed["digest"],
                parsed["sheet_name"],
                parsed["record_count"],
                inserted,
                updated,
                skipped,
                parsed["warning_count"],
                conflict_mode,
                backup_path,
                session["user_id"],
            ),
        )
        db.commit()
    except Exception:
        db.rollback()
        raise

    _remove_import_file(token)
    session.pop("ipc_import", None)
    result = {
        "filename": state.get("filename") or parsed["filename"],
        "source_rows": parsed["record_count"],
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "warning_count": parsed["warning_count"],
        "backup_name": os.path.basename(backup_path),
    }
    return render_template("veri_aktar.html", preview=None, result=result)

# ---------------------------------------------------------------- şirket işlemleri
@app.route("/sirket-ekle", methods=["GET", "POST"])
@login_required
def sirket_ekle():
    if request.method == "POST":
        data, errors = collect_form(request.form)
        prepared, attachment_error = _prepare_attachments(
            request.files.getlist("attachments")
        )
        if errors or attachment_error:
            if errors:
                flash(f"Formda {len(errors)} hata bulundu. Lütfen işaretli alanları düzeltiniz.", "danger")
            form_data = form_data_for_render(request.form)
            form_data["defter_sira_no"] = session["user_name"]
            return render_template("sirket_form.html", sections=form_sections(),
                               data=form_data, errors=errors, mode="ekle",
                               attachments=[],
                               attachment_error=attachment_error)
        data["defter_sira_no"] = session["user_name"]
        db = get_db()
        saved_names = []
        try:
            if session["role"] == "admin":
                result = apply_company_action(
                    "ekle", None, json.dumps(data, ensure_ascii=False),
                    session["user_name"],
                )
                saved_names = _save_prepared_attachments(
                    prepared, company_sira_no=result["sira_no"]
                )
            else:
                cursor = db.execute(
                    """INSERT INTO pending_actions(action_type,data,requested_by)
                       VALUES('ekle',?,?)""",
                    (json.dumps(data, ensure_ascii=False), session["user_id"]),
                )
                saved_names = _save_prepared_attachments(
                    prepared, pending_action_id=cursor.lastrowid
                )
                sync_admin_approval_notifications()
            db.commit()
        except Exception:
            db.rollback()
            _delete_stored_files(saved_names)
            app.logger.exception("Şirket kaydı ve ek dosyaları kaydedilemedi.")
            flash("Kayıt veya ek dosyalar kaydedilemedi. Lütfen tekrar deneyiniz.", "danger")
            return render_template(
                "sirket_form.html", sections=form_sections(),
                data=form_data_for_render(request.form),
                errors={}, mode="ekle", attachments=[],
                attachment_error="Dosyalar kaydedilirken beklenmeyen bir hata oluştu.",
            )

        if session["role"] == "admin":
            flash("İPC kaydı başarıyla oluşturuldu.", "success")
            return redirect(url_for("sirket_oku", sira_no=result["sira_no"]))
        else:
            flash("Kaydınız yönetici onayına gönderildi. Onaylandığında bildirim alacaksınız.", "info")
        return redirect(url_for("sirketler"))
    return render_template("sirket_form.html", sections=form_sections(),
                           data={}, errors={}, mode="ekle", attachments=[],
                           attachment_error=None)

def _find_company(sira_no):
    r = get_db().execute("SELECT * FROM companies WHERE sira_no=?", (sira_no,)).fetchone()
    if not r:
        return None
    c = dict(r)
    c["data"] = json.loads(c["data"])
    c["kalan"] = kalan_sure(c["data"])
    c["status"] = company_status(c)
    c["attachments"] = _attachments_for_company(c["sira_no"])
    return c

def _detail_sections():
    fields = get_fields()
    by_key = {field["field_key"]: field for field in fields}
    used = {"sira_no"}
    sections = []
    for title, description, keys in DETAIL_SECTION_KEYS:
        section_fields = [by_key[key] for key in keys if key in by_key]
        used.update(field["field_key"] for field in section_fields)
        sections.append({
            "title": title,
            "description": description,
            "fields": section_fields,
        })
    remaining = [
        field for field in fields
        if field["field_key"] not in used
    ]
    if remaining:
        sections.append({
            "title": "Diğer Bilgiler",
            "description": "Forma sonradan eklenen veya özel olarak yapılandırılan alanlar.",
            "fields": remaining,
        })
    return sections

@app.route("/sirket-oku", methods=["GET", "POST"])
@login_required
def sirket_oku():
    requested_sira = (
        request.form.get("sira_no", "0")
        if request.method == "POST"
        else request.args.get("sira_no")
    )
    if not requested_sira:
        return redirect(url_for("sirketler"))
    company = _find_company(requested_sira)
    if not company:
        flash("Açmak istediğiniz İPC kaydı bulunamadı.", "danger")
        return redirect(url_for("sirketler"))
    return render_template(
        "sirket_oku.html",
        company=company,
        sections=_detail_sections(),
    )

@app.route("/ek/<int:attachment_id>")
@login_required
def attachment_file(attachment_id):
    row = get_db().execute(
        """SELECT a.*, p.requested_by
           FROM attachments a
           LEFT JOIN pending_actions p ON p.id=a.pending_action_id
           WHERE a.id=?""",
        (attachment_id,),
    ).fetchone()
    if not row:
        abort(404)
    can_access = (
        row["company_sira_no"] is not None
        or session.get("role") == "admin"
        or row["requested_by"] == session.get("user_id")
    )
    if not can_access:
        abort(403)

    path = os.path.join(UPLOAD_DIR, row["stored_name"])
    if not os.path.isfile(path):
        abort(404)
    extension = os.path.splitext(row["stored_name"])[1].lower()
    show_inline = (
        request.args.get("goruntule") == "1"
        and extension in INLINE_ATTACHMENT_EXTENSIONS
    )
    response = send_from_directory(
        UPLOAD_DIR,
        row["stored_name"],
        as_attachment=not show_inline,
        download_name=row["original_name"],
        mimetype=row["mime_type"] or "application/octet-stream",
        conditional=True,
    )
    response.headers["X-Content-Type-Options"] = "nosniff"
    if show_inline:
        response.headers["Content-Security-Policy"] = "sandbox"
    return response

@app.route("/ek/<int:attachment_id>/sil", methods=["POST"])
@login_required
@admin_required
def attachment_delete(attachment_id):
    db = get_db()
    row = db.execute(
        "SELECT * FROM attachments WHERE id=?", (attachment_id,)
    ).fetchone()
    if not row:
        flash("Ek dosya bulunamadı.", "danger")
        return redirect(url_for("sirketler"))
    sira_no = row["company_sira_no"]
    db.execute("DELETE FROM attachments WHERE id=?", (attachment_id,))
    db.commit()
    _delete_stored_files([row["stored_name"]])
    flash(f"“{row['original_name']}” ek dosyası silindi.", "warning")
    if sira_no:
        return redirect(url_for("sirket_oku", sira_no=sira_no))
    return redirect(url_for("onaylar"))

@app.route("/sirket-duzenle", methods=["GET", "POST"])
@login_required
def sirket_duzenle():
    step = request.form.get("step", "sor")
    if request.method == "GET":
        requested_sira = request.args.get("sira_no")
        if not requested_sira:
            return redirect(url_for("sirketler"))
        c = _find_company(requested_sira)
        if not c:
            flash("Düzenlemek istediğiniz İPC kaydı bulunamadı.", "danger")
            return redirect(url_for("sirketler"))
        return render_template(
            "sirket_form.html", sections=form_sections(),
            data=c["data"], errors={}, mode="duzenle",
            sira_no=c["sira_no"], attachments=c["attachments"],
            attachment_error=None,
        )
    if request.method == "POST" and step == "sor":
        c = _find_company(request.form.get("sira_no", "0"))
        if not c:
            flash("Düzenlemek istediğiniz İPC kaydı bulunamadı.", "danger")
            return render_template("sira_sor.html", islem="duzenle")
        return render_template("sirket_form.html", sections=form_sections(),
                               data=c["data"], errors={}, mode="duzenle",
                               sira_no=c["sira_no"],
                               attachments=c["attachments"],
                               attachment_error=None)
    if request.method == "POST" and step == "kaydet":
        sira_no = int(request.form.get("sira_no"))
        c = _find_company(sira_no)
        if not c:
            abort(404)
        data, errors = collect_form(request.form)
        prepared, attachment_error = _prepare_attachments(
            request.files.getlist("attachments")
        )
        if errors or attachment_error:
            if errors:
                flash(f"Formda {len(errors)} hata bulundu. Lütfen işaretli alanları düzeltiniz.", "danger")
            form_data = form_data_for_render(request.form)
            form_data["defter_sira_no"] = (
                c["data"].get("defter_sira_no")
                or session["user_name"]
            )
            return render_template("sirket_form.html", sections=form_sections(),
                                   data=form_data, errors=errors,
                                   mode="duzenle", sira_no=sira_no,
                                   attachments=c["attachments"],
                                   attachment_error=attachment_error)
        data["defter_sira_no"] = (
            c["data"].get("defter_sira_no")
            or session["user_name"]
        )
        payload = json.dumps(data, ensure_ascii=False)
        db = get_db()
        saved_names = []
        try:
            if session["role"] == "admin":
                apply_company_action(
                    "duzenle", sira_no, payload, session["user_name"]
                )
                saved_names = _save_prepared_attachments(
                    prepared, company_sira_no=sira_no
                )
            else:
                cursor = db.execute(
                    """INSERT INTO pending_actions
                       (action_type,sira_no,data,requested_by)
                       VALUES('duzenle',?,?,?)""",
                    (sira_no, payload, session["user_id"]),
                )
                saved_names = _save_prepared_attachments(
                    prepared, pending_action_id=cursor.lastrowid
                )
                sync_admin_approval_notifications()
            db.commit()
        except Exception:
            db.rollback()
            _delete_stored_files(saved_names)
            app.logger.exception("Şirket düzenlemesi ve ek dosyaları kaydedilemedi.")
            flash("Düzenleme veya ek dosyalar kaydedilemedi. Lütfen tekrar deneyiniz.", "danger")
            return render_template(
                "sirket_form.html", sections=form_sections(),
                data=form_data_for_render(request.form),
                errors={}, mode="duzenle", sira_no=sira_no,
                attachments=c["attachments"],
                attachment_error="Dosyalar kaydedilirken beklenmeyen bir hata oluştu.",
            )

        if session["role"] == "admin":
            flash(f"Sıra No {sira_no} kaydı güncellendi.", "success")
        else:
            flash("Düzenleme talebiniz yönetici onayına gönderildi.", "info")
        return redirect(url_for("sirket_oku", sira_no=sira_no))
    return redirect(url_for("sirketler"))

@app.route("/sirket-sil", methods=["GET", "POST"])
@login_required
def sirket_sil():
    if request.method == "POST":
        sira_no = request.form.get("sira_no", "0")
        c = _find_company(sira_no)
        if not c:
            flash("Silmek istediğiniz İPC kaydı bulunamadı.", "danger")
            return render_template("sira_sor.html", islem="sil")
        db = get_db()
        if session["role"] == "admin":
            result = apply_company_action(
                "sil", c["sira_no"], None, session["user_name"]
            )
            db.commit()
            _delete_stored_files(result["deleted_files"])
            flash(f"Sıra No {c['sira_no']} kaydı silindi.", "warning")
        else:
            db.execute("""INSERT INTO pending_actions(action_type,sira_no,requested_by)
                          VALUES('sil',?,?)""", (c["sira_no"], session["user_id"]))
            sync_admin_approval_notifications()
            db.commit()
            flash("Silme talebiniz yönetici onayına gönderildi.", "info")
        return redirect(url_for("sirketler"))
    return redirect(url_for("sirketler"))

# ---------------------------------------------------------------- liste & filtre
def tr_kucuk(s):
    """Türkçe'ye uygun küçük harfe çevirme (I→ı, İ→i)."""
    return str(s).replace("I", "ı").replace("İ", "i").lower()

# Aralık ifadeleri:  "30.10.2024-02.04.2026" · "01.01.2026-" · "-31.12.2025"
ARALIK_TARIH = re.compile(
    r"^\s*(\d{1,2}\.\d{1,2}\.\d{2,4})?\s*-\s*(\d{1,2}\.\d{1,2}\.\d{2,4})?\s*$")
# Aralık ifadeleri:  "2025-2026" · "2025-" · "16000-51000"
ARALIK_SAYI = re.compile(
    r"^(-?\d+(?:[.,]\d+)?)-(-?\d+(?:[.,]\d+)?)?$")

def sayiya_cevir(s):
    try:
        return float(str(s).replace(" ", "").replace(",", "."))
    except (ValueError, TypeError):
        return None

def filtre_esles(deger, arama, datatype):
    """Bir hücre değerinin filtre ifadesine uyup uymadığını döndürür.

    Tarih alanları : tek tarih ("2.4.2026") veya aralık ("30.10.2024-2.04.2026").
    Sayı alanları  : tek sayı ("2025") veya aralık ("2025-2026"). Boşluklu para
                     girişi ("2 345 234") burada da kabul edilir.
    Diğer alanlar  : Türkçe duyarlı, büyük/küçük harf ayırmayan içerik araması.
    """
    arama = (arama or "").strip()
    if not arama:
        return True

    # ---- tarih alanları ----
    if datatype.startswith("TARİH") or datatype.startswith("DATETIME"):
        m = ARALIK_TARIH.match(arama)
        if m and (m.group(1) or m.group(2)):
            d = parse_tr_date(deger)
            if d is None:
                return False
            bas = parse_tr_date(m.group(1)) if m.group(1) else None
            bit = parse_tr_date(m.group(2)) if m.group(2) else None
            if m.group(1) and bas is None:
                return False
            if m.group(2) and bit is None:
                return False
            if bas and d < bas:
                return False
            if bit and d > bit:
                return False
            return True
        tek = parse_tr_date(arama)
        if tek:
            return parse_tr_date(deger) == tek
        return tr_kucuk(arama) in tr_kucuk(deger)     # yarım yazılmışsa metin araması

    # ---- sayı alanları (Sıra No, Yıl, Kalan Süre, tutarlar) ----
    if datatype == "INT" or datatype in PARA_TIPLERI:
        temiz = arama.replace(" ", "")
        n = sayiya_cevir(deger)
        m = ARALIK_SAYI.match(temiz)
        if m:
            if n is None:
                return False
            bas = sayiya_cevir(m.group(1))
            bit = sayiya_cevir(m.group(2)) if m.group(2) else None
            if bas is not None and n < bas:
                return False
            if bit is not None and n > bit:
                return False
            return True
        tek = sayiya_cevir(temiz)
        if tek is not None:
            return n is not None and n == tek
        return tr_kucuk(arama) in tr_kucuk(deger)

    # ---- metin alanları ----
    return tr_kucuk(arama) in tr_kucuk(deger)

def _has_positive_amount(data, *keys):
    for key in keys:
        try:
            if float(str(data.get(key) or "0").replace(",", ".")) > 0:
                return True
        except ValueError:
            continue
    return False

def company_status(company):
    data = company["data"]
    source_status = tr_kucuk(data.get("kaynak_durumu") or "")
    source_status = source_status.replace("ı̇", "i")
    unpaid_words = ("yapılmadı", "alınmadı", "alınamadı")
    paid_words = ("yapıldı", "alındı", "yapılmış", "ödendi")
    source_paid = (
        any(word in source_status for word in paid_words)
        and not any(word in source_status for word in unpaid_words)
    )
    if "iptal" in source_status or "iptal" in tr_kucuk(data.get("gonderim_turu") or ""):
        return {"key": "iptal", "label": "İptal edildi", "tone": "neutral"}
    if (
        data.get("odeme_tarihi")
        or data.get("vergi_tahsil_tarihi")
        or source_paid
        or _has_positive_amount(
            data, "dogrudan_yatirilan", "itiraz_lehine_tutar",
            "vergi_tahsil_tutari",
        )
    ):
        return {"key": "odendi", "label": "Ödendi", "tone": "success"}
    if data.get("itiraz_sonucu"):
        return {"key": "itiraz", "label": "İtiraz sürecinde", "tone": "info"}
    if data.get("vergi_bildirim_tarihi") or data.get("itiraz_vergi_bildirim_tarihi"):
        return {
            "key": "vergi-bildirildi",
            "label": "Vergi dairesine bildirildi",
            "tone": "info",
        }
    if data.get("kesinlesme_tarihi"):
        return {"key": "kesinlesti", "label": "Kesinleşti", "tone": "info"}
    if company["kalan"] is not None:
        if company["kalan"] < 0:
            return {"key": "sure-doldu", "label": "Süre doldu", "tone": "danger"}
        return {
            "key": "sure-isliyor",
            "label": f"Süre işliyor · {company['kalan']} gün",
            "tone": "warning",
        }
    if data.get("ceza_onay_tarihi"):
        return {"key": "teblig-bekliyor", "label": "Tebliğ bekliyor", "tone": "warning"}
    return {"key": "yeni", "label": "Yeni kayıt", "tone": "neutral"}

def _load_companies():
    companies = []
    for r in get_db().execute("SELECT * FROM companies ORDER BY sira_no DESC"):
        c = dict(r)
        c["data"] = json.loads(c["data"])
        c["kalan"] = kalan_sure(c["data"])
        c["status"] = company_status(c)
        companies.append(c)
    return companies

def _apply_company_filters(companies, args, include_status=True):
    query = tr_kucuk(args.get("q", "").strip())
    if query:
        companies = [
            c for c in companies
            if query in tr_kucuk(c["data"].get("tablet_tutanak_no", ""))
            or query in tr_kucuk(c["data"].get("cezanin_muhatabi", ""))
        ]

    year = args.get("yil", "").strip()
    if year:
        companies = [
            c for c in companies if str(c["data"].get("yil") or "") == year
        ]

    year_from = args.get("yil_baslangic", "").strip()
    year_to = args.get("yil_bitis", "").strip()
    try:
        start_year = int(year_from) if year_from else None
    except ValueError:
        start_year = None
    try:
        end_year = int(year_to) if year_to else None
    except ValueError:
        end_year = None
    if start_year is not None and end_year is not None and start_year > end_year:
        start_year, end_year = end_year, start_year
    if start_year is not None or end_year is not None:
        def within_year_range(company):
            try:
                company_year = int(str(company["data"].get("yil") or "").strip())
            except (TypeError, ValueError):
                return False
            return (
                (start_year is None or company_year >= start_year)
                and (end_year is None or company_year <= end_year)
            )

        companies = [company for company in companies if within_year_range(company)]

    legal_basis = args.get("hukuki_dayanak", "").strip()
    if legal_basis:
        companies = [
            c for c in companies
            if str(c["data"].get("hukuki_dayanak") or "") == legal_basis
        ]

    if include_status:
        status = args.get("durum", "").strip()
        if status:
            companies = [c for c in companies if c["status"]["key"] == status]

    tipler = {f["field_key"]: f["datatype"] for f in get_fields()}
    fkeys = args.getlist("f_field")
    fvals = args.getlist("f_value")
    filters = [(k, v.strip()) for k, v in zip(fkeys, fvals) if k and v.strip()]
    for key, val in filters:
        dt = tipler.get(key, "TEXT")
        companies = [
            c for c in companies
            if filtre_esles(company_cell(c, key), val, dt)
        ]
    return companies, filters

def filtered_companies(args):
    return _apply_company_filters(_load_companies(), args)

def report_companies(args):
    companies, filters = filtered_companies(args)

    def chronological_key(company):
        try:
            year = int(str(company["data"].get("yil") or "").strip())
        except (TypeError, ValueError):
            year = 9999
        return year, int(company["sira_no"])

    return sorted(companies, key=chronological_key), filters

def company_cell(c, key):
    if key == "sira_no":
        return str(c["sira_no"])
    if key == "kalan_sure":
        return "" if c["kalan"] is None else str(c["kalan"])
    return str(c["data"].get(key, "") or "")

@app.template_filter("money_tr")
def format_money_tr(value):
    if value in (None, ""):
        return "—"
    try:
        number = float(str(value).replace(" ", "").replace(",", "."))
    except (TypeError, ValueError):
        return str(value)
    formatted = f"{number:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"{formatted} TL"

@app.route("/sirketler")
@login_required
def sirketler():
    all_companies = _load_companies()
    filtered, _ = _apply_company_filters(all_companies, request.args)
    status_base, _ = _apply_company_filters(
        all_companies, request.args, include_status=False
    )
    status_counts = {
        "": len(status_base),
        "sure-isliyor": sum(
            c["status"]["key"] == "sure-isliyor" for c in status_base
        ),
        "sure-doldu": sum(
            c["status"]["key"] == "sure-doldu" for c in status_base
        ),
        "odendi": sum(c["status"]["key"] == "odendi" for c in status_base),
    }

    try:
        per_page = int(request.args.get("sayfa_boyutu", 25))
    except (TypeError, ValueError):
        per_page = 25
    if per_page not in (25, 50, 100):
        per_page = 25
    total = len(filtered)
    page_count = max(1, (total + per_page - 1) // per_page)
    try:
        page = int(request.args.get("sayfa", 1))
    except (TypeError, ValueError):
        page = 1
    page = min(max(page, 1), page_count)
    start = (page - 1) * per_page
    companies = filtered[start:start + per_page]

    def page_url(target_page):
        params = request.args.to_dict(flat=True)
        params["sayfa"] = target_page
        return url_for("sirketler", **params)

    page_numbers = sorted({
        number for number in (
            1, page_count, page - 2, page - 1, page, page + 1, page + 2
        ) if 1 <= number <= page_count
    })
    pagination = {
        "page": page,
        "page_count": page_count,
        "total": total,
        "start": start + 1 if total else 0,
        "end": min(start + per_page, total),
        "numbers": [
            {"number": number, "url": page_url(number)}
            for number in page_numbers
        ],
        "previous_url": page_url(page - 1) if page > 1 else None,
        "next_url": page_url(page + 1) if page < page_count else None,
    }
    years = sorted({
        str(c["data"].get("yil"))
        for c in all_companies if c["data"].get("yil")
    }, reverse=True)
    legal_bases = sorted({
        str(c["data"].get("hukuki_dayanak"))
        for c in all_companies if c["data"].get("hukuki_dayanak")
    }, key=tr_kucuk)
    return render_template(
        "sirketler.html",
        companies=companies,
        fields=get_fields(),
        company_cell=company_cell,
        pagination=pagination,
        status_counts=status_counts,
        years=years,
        legal_bases=legal_bases,
        per_page=per_page,
    )

# ---------------------------------------------------------------- dışa aktarma
def build_excel(companies, fields, title):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    wb.properties.title = title
    wb.properties.creator = "Kocaeli İl Müdürlüğü İPC Yönetim Sistemi"
    ws = wb.active
    ws.title = "İPC"
    last_column_letter = get_column_letter(len(fields))
    # Keep the identity block within the first printed horizontal page.  The
    # full report can contain many wide columns, so merging across all of them
    # would place the title far outside the initial Excel viewport.
    header_merge_letter = get_column_letter(min(len(fields), 6))

    years = []
    for company in companies:
        try:
            years.append(int(str(company["data"].get("yil") or "").strip()))
        except (TypeError, ValueError):
            continue
    if years:
        period = str(min(years)) if min(years) == max(years) else f"{min(years)}–{max(years)}"
    else:
        period = "Belirtilmemiş"
    total_amount = sum(
        amount_value(company["data"].get("ceza_tutari"))
        for company in companies
    )

    ws.merge_cells(f"A1:{header_merge_letter}1")
    ws.merge_cells(f"A2:{header_merge_letter}2")
    ws["A1"] = f"KOCAELİ İL MÜDÜRLÜĞÜ | {title}"
    ws["A2"] = (
        f"Dönem: {period}  •  Oluşturma Tarihi: {date.today().strftime('%d.%m.%Y')}"
        f"  •  Kayıt Sayısı: {len(companies)}"
        f"  •  Toplam Ceza Tutarı: {format_money_tr(total_amount)}"
    )
    ws.append([f["label"] for f in fields])

    preferred_widths = {
        "sira_no": 11,
        "defter_sira_no": 19,
        "yil": 9,
        "il_adi": 34,
        "hukuki_dayanak": 18,
        "olcu_aleti_sayisi": 16,
        "olcu_aleti_cinsi": 42,
        "kaynak_notu": 42,
        "tablet_tutanak_no": 23,
        "cezanin_muhatabi": 36,
        "ceza_onay_tarihi": 21,
        "ceza_tutari": 21,
        "gonderim_turu": 17,
        "teblig_tarihi": 21,
        "dogrudan_yatirilan": 24,
        "odeme_tarihi": 18,
        "kesinlesme_tarihi": 22,
        "kalan_sure": 15,
        "vergi_bildirim_tarihi": 27,
        "vergi_dairesi": 27,
        "itiraz_sonucu": 30,
        "itiraz_lehine_tutar": 26,
        "itiraz_vergi_bildirim_tarihi": 29,
        "vergi_tahsil_tutari": 25,
        "vergi_tahsil_tarihi": 24,
        "tespit_kurum": 23,
        "ebys": 28,
        "kaynak_durumu": 30,
    }
    burgundy_fill = PatternFill("solid", fgColor="8C1D2F")
    metadata_fill = PatternFill("solid", fgColor="F2E8EA")
    alternate_fill = PatternFill("solid", fgColor="F4F5F7")
    thin_gray = Side(style="thin", color="D5D9DE")
    cell_border = Border(
        left=thin_gray,
        right=thin_gray,
        top=thin_gray,
        bottom=thin_gray,
    )

    widths = {}
    for column_index, field in enumerate(fields, start=1):
        letter = get_column_letter(column_index)
        width = preferred_widths.get(field["field_key"])
        if width is None:
            longest_value = max(
                [
                    len(str(field["label"] or "")),
                    *[
                        len(company_cell(company, field["field_key"]))
                        for company in companies
                    ],
                ]
            )
            width = min(max(longest_value + 2, 14), 42)
        widths[column_index] = width
        ws.column_dimensions[letter].width = width

    for column_index in range(1, len(fields) + 1):
        ws.cell(row=1, column=column_index).fill = burgundy_fill
        ws.cell(row=2, column=column_index).fill = metadata_fill
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].font = Font(name="Calibri", size=10, bold=True, color="6A1423")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 24

    for cell in ws[3]:
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = burgundy_fill
        cell.border = cell_border
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    header_lines = max(
        (
            max(
                1,
                (len(str(field["label"] or "")) + int(widths[index]) - 3)
                // max(int(widths[index]) - 2, 1),
            )
            for index, field in enumerate(fields, start=1)
        ),
        default=1,
    )
    ws.row_dimensions[3].height = min(max(36, header_lines * 15), 96)

    for c in companies:
        ws.append([company_cell(c, f["field_key"]) for f in fields])

    for row_index in range(4, ws.max_row + 1):
        maximum_lines = 1
        for column_index, field in enumerate(fields, start=1):
            cell = ws.cell(row=row_index, column=column_index)
            value_length = len(str(cell.value or ""))
            maximum_lines = max(
                maximum_lines,
                max(
                    1,
                    (value_length + int(widths[column_index]) - 1)
                    // max(int(widths[column_index]), 1),
                ),
            )
            horizontal = "left"
            if (
                field["datatype"] == "INT"
                or field["datatype"].startswith("TARİH")
                or field["datatype"].startswith("DATETIME")
            ):
                horizontal = "center"
            elif field["datatype"] in PARA_TIPLERI:
                horizontal = "right"
            cell.font = Font(name="Calibri", size=10, color="20242A")
            cell.border = cell_border
            cell.alignment = Alignment(
                horizontal=horizontal,
                vertical="top",
                wrap_text=True,
            )
            if (row_index - 4) % 2 == 0:
                cell.fill = alternate_fill
        ws.row_dimensions[row_index].height = min(max(20, maximum_lines * 15), 60)

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 85
    ws.print_title_rows = "1:3"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.scale = 85
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def build_pdf(companies, fields, title, rapor=False):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, Image as RLImage,
                                    PageBreak)
    from reportlab.lib.styles import ParagraphStyle
    from xml.sax.saxutils import escape
    fdir = os.path.join(BASE_DIR, "fonts")
    pdfmetrics.registerFont(TTFont("DejaVu", os.path.join(fdir, "DejaVuSans.ttf")))
    pdfmetrics.registerFont(TTFont("DejaVu-Bold", os.path.join(fdir, "DejaVuSans-Bold.ttf")))
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=10*mm, rightMargin=10*mm,
                            topMargin=10*mm, bottomMargin=12*mm)
    stil_baslik = ParagraphStyle("b", fontName="DejaVu-Bold", fontSize=13,
                                 alignment=1, spaceAfter=2)
    stil_alt = ParagraphStyle("a", fontName="DejaVu", fontSize=9, alignment=1,
                              textColor=colors.HexColor("#555555"))
    stil_hucre = ParagraphStyle("h", fontName="DejaVu", fontSize=7, leading=8.5)
    stil_bas = ParagraphStyle("hb", fontName="DejaVu-Bold", fontSize=7,
                              leading=8.5, textColor=colors.white)
    stil_kayit = ParagraphStyle(
        "record-title", fontName="DejaVu-Bold", fontSize=10, leading=12,
        textColor=colors.white,
    )
    stil_bolum = ParagraphStyle(
        "record-section", fontName="DejaVu-Bold", fontSize=8, leading=10,
        textColor=colors.HexColor("#6A1423"),
    )
    stil_etiket = ParagraphStyle(
        "record-label", fontName="DejaVu-Bold", fontSize=6.7, leading=8.2,
        textColor=colors.HexColor("#4A515A"),
    )
    stil_deger = ParagraphStyle(
        "record-value", fontName="DejaVu", fontSize=7.5, leading=9.3,
        textColor=colors.HexColor("#20242A"),
    )
    story = []
    logo_path = os.path.join(FRONTEND, "static", "img", "logo.png")
    if os.path.exists(logo_path):
        img = RLImage(logo_path, width=60*mm, height=16*mm)
        img.hAlign = "CENTER"
        story.append(img)
        story.append(Spacer(1, 3*mm))
    story.append(Paragraph("T.C. SANAYİ VE TEKNOLOJİ BAKANLIĞI", stil_baslik))
    story.append(Paragraph("Kocaeli İl Müdürlüğü", stil_alt))
    story.append(Spacer(1, 2*mm))
    story.append(Paragraph(title, stil_baslik))
    story.append(Paragraph(
        f"Oluşturma Tarihi: {date.today().strftime('%d.%m.%Y')}  •  Kayıt Sayısı: {len(companies)}",
        stil_alt))
    if rapor:
        toplam = 0.0
        for c in companies:
            try:
                toplam += float(c["data"].get("ceza_tutari") or 0)
            except ValueError:
                pass
        story.append(Paragraph(
            f"Toplam Ceza Tutarı: {format_money_tr(toplam)}",
            stil_alt,
        ))
    story.append(Spacer(1, 4*mm))

    detailed_layout = len(fields) > 12
    if detailed_layout:
        story.append(Paragraph(
            "Okunabilirlik için boş alanlar PDF'de gösterilmez; Excel çıktısı "
            "28 sütunun tamamını korur.",
            stil_alt,
        ))
        story.append(Spacer(1, 2*mm))
        fields_by_key = {field["field_key"]: field for field in fields}
        section_definitions = [
            (section_title, section_keys)
            for section_title, _description, section_keys
            in DETAIL_SECTION_KEYS
        ]
        assigned_keys = {"sira_no"} | {
            key
            for _section_title, section_keys in section_definitions
            for key in section_keys
        }
        remaining_keys = [
            field["field_key"] for field in fields
            if field["field_key"] not in assigned_keys
        ]
        if remaining_keys:
            section_definitions.append(("Diğer Bilgiler", tuple(remaining_keys)))

        label_width = 38 * mm
        value_width = (doc.width - (2 * label_width)) / 2
        column_widths = [label_width, value_width, label_width, value_width]

        if not companies:
            story.append(Paragraph("Raporlanacak kayıt bulunamadı.", stil_hucre))

        for company_index, company in enumerate(companies):
            if company_index:
                story.append(PageBreak())
            record_number = escape(str(company["sira_no"]))
            company_name = escape(
                str(company["data"].get("cezanin_muhatabi") or "Kişi/firma adı bulunmuyor")
            )
            record_header = Table(
                [[Paragraph(f"Sıra No {record_number}  ·  {company_name}", stil_kayit)]],
                colWidths=[doc.width],
            )
            record_header.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#8C1D2F")),
                ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#6A1423")),
                ("LEFTPADDING", (0, 0), (-1, -1), 7),
                ("RIGHTPADDING", (0, 0), (-1, -1), 7),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]))
            story.append(record_header)
            story.append(Spacer(1, 2.5*mm))

            for section_title, section_keys in section_definitions:
                section_values = []
                for key in section_keys:
                    if key not in fields_by_key:
                        continue
                    field = fields_by_key[key]
                    raw_value = company_cell(company, key)
                    if not raw_value and key != "sira_no":
                        continue
                    section_values.append((field, raw_value))
                if not section_values:
                    continue
                rows = [[
                    Paragraph(escape(section_title), stil_bolum), "", "", ""
                ]]
                for field_index in range(0, len(section_values), 2):
                    row = []
                    for field, raw_value in section_values[field_index:field_index + 2]:
                        if field["datatype"] in PARA_TIPLERI and raw_value:
                            raw_value = format_money_tr(raw_value)
                        row.extend([
                            Paragraph(escape(str(field["label"])), stil_etiket),
                            Paragraph(escape(str(raw_value or "—")), stil_deger),
                        ])
                    while len(row) < 4:
                        row.extend(["", ""])
                    rows.append(row)
                section_table = Table(
                    rows,
                    colWidths=column_widths,
                    repeatRows=1,
                    splitByRow=1,
                )
                section_table.setStyle(TableStyle([
                    ("SPAN", (0, 0), (-1, 0)),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F2E8EA")),
                    ("BOX", (0, 0), (-1, -1), 0.45, colors.HexColor("#AAB0B7")),
                    ("INNERGRID", (0, 1), (-1, -1), 0.3, colors.HexColor("#D5D9DE")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("BACKGROUND", (0, 1), (0, -1), colors.HexColor("#F7F8F9")),
                    ("BACKGROUND", (2, 1), (2, -1), colors.HexColor("#F7F8F9")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 5),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]))
                story.append(section_table)
                story.append(Spacer(1, 2*mm))
    else:
        header = [Paragraph(f["label"], stil_bas) for f in fields]
        rows = [header]
        for c in companies:
            rows.append([
                Paragraph(company_cell(c, f["field_key"]), stil_hucre)
                for f in fields
            ])
        table = Table(rows, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#8C1D2F")),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#999999")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.white, colors.HexColor("#F5EFE6")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(table)
    if rapor and not detailed_layout:
        story.append(Spacer(1, 10*mm))
        stil_imza = ParagraphStyle("i", fontName="DejaVu", fontSize=9, alignment=2)
        story.append(Paragraph("Düzenleyen: " + session.get("user_name", ""), stil_imza))
        story.append(Paragraph("İmza: ______________________", stil_imza))

    report_actor = session.get("user_name", "") if rapor else ""

    def add_page_footer(canvas, _document):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor("#D5D9DE"))
        canvas.setLineWidth(0.4)
        canvas.line(10*mm, 9*mm, landscape(A4)[0] - 10*mm, 9*mm)
        canvas.setFont("DejaVu", 7)
        canvas.setFillColor(colors.HexColor("#626A74"))
        canvas.drawString(10*mm, 5.5*mm, title)
        if detailed_layout and report_actor:
            canvas.drawCentredString(
                landscape(A4)[0] / 2,
                5.5*mm,
                f"Düzenleyen: {report_actor}  •  İmza: ____________________",
            )
        canvas.drawRightString(
            landscape(A4)[0] - 10*mm,
            5.5*mm,
            f"Sayfa {canvas.getPageNumber()}",
        )
        canvas.restoreState()

    doc.build(
        story,
        onFirstPage=add_page_footer,
        onLaterPages=add_page_footer,
    )
    buf.seek(0)
    return buf

@app.route("/disa-aktar/<fmt>")
@login_required
def disa_aktar(fmt):
    companies, _ = filtered_companies(request.args)
    fields = get_fields()
    if fmt == "excel":
        buf = build_excel(companies, fields, "İdari Para Cezaları")
        return send_file(buf, as_attachment=True, download_name="idari_para_cezalari.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    buf = build_pdf(companies, fields, "İdari Para Cezaları Listesi")
    return send_file(buf, as_attachment=True, download_name="idari_para_cezalari.pdf",
                     mimetype="application/pdf")

# ---------------------------------------------------------------- rapor
def _report_fields(args):
    fields = get_fields()
    by_key = {field["field_key"]: field for field in fields}
    selected = [
        key for key in args.getlist("cols") if key in by_key
    ]
    report_type = args.get("rapor_turu", "ozet")
    if report_type not in REPORT_PRESETS:
        report_type = "ozet"
    keys = selected or REPORT_PRESETS[report_type]["keys"]
    return report_type, [by_key[key] for key in keys if key in by_key]

def _report_summary(companies):
    total_amount = 0.0
    for company in companies:
        try:
            total_amount += float(company["data"].get("ceza_tutari") or 0)
        except (TypeError, ValueError):
            continue
    return {
        "record_count": len(companies),
        "total_amount": total_amount,
        "active_count": sum(
            company["status"]["key"] == "sure-isliyor"
            for company in companies
        ),
        "expired_count": sum(
            company["status"]["key"] == "sure-doldu"
            for company in companies
        ),
        "paid_count": sum(
            company["status"]["key"] == "odendi"
            for company in companies
        ),
    }

@app.route("/rapor", methods=["GET"])
@login_required
def rapor():
    all_companies = _load_companies()
    companies, filters = report_companies(request.args)
    report_type, report_fields = _report_fields(request.args)
    preview = request.args.get("onizle") == "1"
    years = sorted({
        str(company["data"].get("yil"))
        for company in all_companies if company["data"].get("yil")
    }, reverse=True)
    legal_bases = sorted({
        str(company["data"].get("hukuki_dayanak"))
        for company in all_companies if company["data"].get("hukuki_dayanak")
    }, key=tr_kucuk)
    seen_statuses = {
        company["status"]["key"] for company in all_companies
    }
    statuses = [
        {"key": key, "label": label}
        for key, label in STATUS_FILTER_LABELS.items()
        if key in seen_statuses
    ]
    return render_template(
        "rapor.html",
        fields=get_fields(),
        companies=companies if preview else [],
        filtered_count=len(companies),
        rapor_fields=report_fields,
        report_type=report_type,
        report_presets=REPORT_PRESETS,
        preview=preview,
        summary=_report_summary(companies),
        years=years,
        legal_bases=legal_bases,
        statuses=statuses,
        company_cell=company_cell,
        bugun=date.today().strftime("%d.%m.%Y"),
    )

@app.route("/rapor/pdf")
@login_required
def rapor_pdf():
    companies, _ = report_companies(request.args)
    report_type, fields = _report_fields(request.args)
    title = REPORT_PRESETS[report_type]["title"]
    buf = build_pdf(companies, fields, title, rapor=True)
    return send_file(buf, as_attachment=True, download_name="ipc_raporu.pdf",
                     mimetype="application/pdf")

@app.route("/rapor/excel")
@login_required
def rapor_excel():
    companies, _ = report_companies(request.args)
    report_type, fields = _report_fields(request.args)
    title = REPORT_PRESETS[report_type]["title"]
    buf = build_excel(companies, fields, title)
    return send_file(buf, as_attachment=True, download_name="ipc_raporu.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ---------------------------------------------------------------- onaylar (admin)
@app.route("/onaylar")
@login_required
@admin_required
def onaylar():
    db = get_db()
    rows = db.execute("""SELECT p.*, COALESCE(u.name,'Silinmiş kullanıcı') requester
                         FROM pending_actions p
                         LEFT JOIN users u ON u.id=p.requested_by
                         ORDER BY CASE p.status WHEN 'bekliyor' THEN 0 ELSE 1 END,
                                  p.id DESC LIMIT 100""").fetchall()
    fields = get_fields()
    items = []
    for r in rows:
        d = dict(r)
        d["data"] = json.loads(r["data"]) if r["data"] else None
        d["attachments"] = _attachments_for_pending(d["id"])
        current = None
        if d["sira_no"]:
            current_row = db.execute(
                "SELECT data FROM companies WHERE sira_no=?", (d["sira_no"],)
            ).fetchone()
            current = json.loads(current_row["data"]) if current_row else None
        incoming = d["data"] or {}
        d["company_name"] = (
            incoming.get("cezanin_muhatabi")
            or (current or {}).get("cezanin_muhatabi")
            or "Firma bilgisi bulunmuyor"
        )
        d["changes"] = []
        if d["action_type"] == "duzenle" and current is not None:
            for field in fields:
                key = field["field_key"]
                old_value = str(current.get(key) or "")
                new_value = str(incoming.get(key) or "")
                if old_value != new_value:
                    d["changes"].append(
                        {
                            "label": field["label"],
                            "old": old_value or "—",
                            "new": new_value or "—",
                        }
                    )
        items.append(d)
    return render_template(
        "onaylar.html",
        pending_items=[item for item in items if item["status"] == "bekliyor"],
        history_items=[item for item in items if item["status"] != "bekliyor"],
        fields=fields,
    )

@app.route("/onaylar/<int:pid>/<karar>", methods=["POST"])
@login_required
@admin_required
def onay_karar(pid, karar):
    db = get_db()
    p = db.execute("SELECT * FROM pending_actions WHERE id=? AND status='bekliyor'",
                   (pid,)).fetchone()
    if not p:
        flash("Bekleyen talep bulunamadı.", "danger")
        return redirect(url_for("onaylar"))
    if karar not in ("onayla", "reddet"):
        abort(400)
    decision_reason = request.form.get("decision_reason", "").strip()
    if karar == "reddet" and not decision_reason:
        flash("Talebi reddetmek için kısa bir gerekçe yazınız.", "danger")
        return redirect(url_for("onaylar", talep=pid))
    isim = {"ekle": "ekleme", "duzenle": "düzenleme", "sil": "silme"}[p["action_type"]]
    pending_files = _attachments_for_pending(pid)
    files_to_delete = []
    if karar == "onayla":
        istekci = db.execute("SELECT name FROM users WHERE id=?",
                             (p["requested_by"],)).fetchone()
        result = apply_company_action(
            p["action_type"], p["sira_no"], p["data"],
            istekci["name"] if istekci else "Bilinmiyor",
        )
        if p["action_type"] in ("ekle", "duzenle"):
            db.execute(
                """UPDATE attachments
                   SET company_sira_no=?, pending_action_id=NULL
                   WHERE pending_action_id=?""",
                (result["sira_no"], pid),
            )
        else:
            files_to_delete.extend(item["stored_name"] for item in pending_files)
            db.execute(
                "DELETE FROM attachments WHERE pending_action_id=?", (pid,)
            )
        files_to_delete.extend(result["deleted_files"])
        db.execute(
            """UPDATE pending_actions
               SET status='onaylandi', decided_at=datetime('now','localtime'),
                   decision_reason=NULL WHERE id=?""",
            (pid,),
        )
        action_url = (
            url_for("sirket_oku", sira_no=result["sira_no"])
            if p["action_type"] != "sil" else None
        )
        notify(
            p["requested_by"],
            f"İPC kaydı {isim} talebiniz onaylandı.",
            title="Talebiniz onaylandı",
            kind="approval-result",
            priority="info",
            action_url=action_url,
        )
        flash("Talep onaylandı ve uygulandı.", "success")
    else:
        files_to_delete.extend(item["stored_name"] for item in pending_files)
        db.execute("DELETE FROM attachments WHERE pending_action_id=?", (pid,))
        db.execute(
            """UPDATE pending_actions
               SET status='reddedildi', decided_at=datetime('now','localtime'),
                   decision_reason=? WHERE id=?""",
            (decision_reason, pid),
        )
        notify(
            p["requested_by"],
            f"İPC kaydı {isim} talebiniz reddedildi. Gerekçe: {decision_reason}",
            title="Talebiniz reddedildi",
            kind="approval-result",
            priority="info",
        )
        flash("Talep reddedildi.", "warning")
    sync_admin_approval_notifications(mark_unread=False)
    db.commit()
    _delete_stored_files(files_to_delete)
    return redirect(url_for("onaylar"))

# ---------------------------------------------------------------- bildirim & şifre
@app.route("/bildirimler")
@login_required
def bildirimler():
    db = get_db()
    sync_admin_approval_notifications(mark_unread=False)
    sync_deadline_notifications(session["user_id"])
    db.commit()
    action_rows = db.execute(
        """SELECT * FROM notifications
           WHERE user_id=? AND resolved_at IS NULL
             AND priority IN ('critical','action')
           ORDER BY CASE priority WHEN 'critical' THEN 0 ELSE 1 END, id DESC
           LIMIT 5""",
        (session["user_id"],),
    ).fetchall()
    recent_rows = db.execute(
        """SELECT * FROM notifications
           WHERE user_id=?
             AND (resolved_at IS NOT NULL OR priority NOT IN ('critical','action'))
           ORDER BY id DESC LIMIT 20""",
        (session["user_id"],),
    ).fetchall()
    return render_template(
        "bildirimler.html",
        action_rows=action_rows,
        recent_rows=recent_rows,
    )

@app.route("/bildirimler/<int:nid>/ac")
@login_required
def bildirim_ac(nid):
    db = get_db()
    notification = db.execute(
        "SELECT * FROM notifications WHERE id=? AND user_id=?",
        (nid, session["user_id"]),
    ).fetchone()
    if not notification:
        abort(404)
    db.execute("UPDATE notifications SET is_read=1 WHERE id=?", (nid,))
    db.commit()
    target = notification["action_url"] or url_for("bildirimler")
    if not target.startswith("/") or target.startswith("//"):
        target = url_for("bildirimler")
    return redirect(target)

@app.route("/bildirimler/<int:nid>/okundu", methods=["POST"])
@login_required
def bildirim_okundu(nid):
    db = get_db()
    db.execute(
        "UPDATE notifications SET is_read=1 WHERE id=? AND user_id=?",
        (nid, session["user_id"]),
    )
    db.commit()
    return redirect(url_for("bildirimler"))

@app.route("/bildirimler/tumunu-okundu", methods=["POST"])
@login_required
def bildirim_tumunu_okundu():
    db = get_db()
    db.execute(
        """UPDATE notifications SET is_read=1
           WHERE user_id=? AND resolved_at IS NULL""",
        (session["user_id"],),
    )
    db.commit()
    return redirect(url_for("bildirimler"))

@app.route("/sifre-degistir", methods=["GET", "POST"])
@login_required
def sifre_degistir():
    if request.method == "POST":
        eski = request.form.get("old", "")
        p1, p2 = request.form.get("password", ""), request.form.get("password2", "")
        db = get_db()
        u = db.execute("SELECT * FROM users WHERE id=?", (session["user_id"],)).fetchone()
        if not check_password_hash(u["password_hash"], eski):
            flash("Mevcut şifreniz hatalı.", "danger")
        elif len(p1) < 6:
            flash("Yeni şifre en az 6 karakter olmalıdır.", "danger")
        elif p1 != p2:
            flash("Girdiğiniz yeni şifreler birbiriyle eşleşmiyor.", "danger")
        else:
            db.execute("UPDATE users SET password_hash=? WHERE id=?",
                       (generate_password_hash(p1), session["user_id"]))
            db.commit()
            flash("Şifreniz başarıyla değiştirildi.", "success")
            return redirect(url_for("dashboard"))
    return render_template("sifre.html")

# ---------------------------------------------------------------- çalıştır
init_db()

if __name__ == "__main__":
    print("=" * 60)
    print(" İPC Yönetim Sistemi -> http://127.0.0.1:5000")
    print(f" Yönetici e-postası: {INITIAL_ADMIN_EMAIL}")
    print("=" * 60)
    app.run(
        host=os.environ.get("HOST", "127.0.0.1"),
        port=int(os.environ.get("PORT", "5000")),
        debug=os.environ.get("FLASK_DEBUG") == "1",
    )
