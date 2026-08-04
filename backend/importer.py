# -*- coding: utf-8 -*-
"""Safe, tolerant importer for the office's historical İPC workbook."""

from __future__ import annotations

import hashlib
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel


class ImportWorkbookError(ValueError):
    """Raised when an uploaded file is not a recognizable İPC workbook."""


# Fields that exist in the real 2017–2026 workbook but not in the first website
# prototype. They are added to form_fields only when a real-data import is
# confirmed, so merely previewing a file never changes the application database.
IMPORT_FIELD_DEFINITIONS = [
    ("defter_sira_no", "Sorumlu Personel", "VARCHAR(n)", 0, None, 1),
    ("il_adi", "İlin Adı (Sorumlu Teşkilat)", "VARCHAR(n)", 0, None, 0),
    (
        "vergi_tahsil_tutari",
        "İlgili Vergi Dairesince Tahsil Edildiği Bildirilen Tutar (TL)",
        "DECIMAL",
        0,
        None,
        0,
    ),
    (
        "vergi_tahsil_tarihi",
        "İlgili Vergi Dairesince Tahsil Edildiği Bildirilen Tarih",
        "TARİH (GG.AA.YYYY)",
        0,
        None,
        0,
    ),
    ("kaynak_durumu", "Kaynak Dosyadaki Durum", "TEXT", 0, None, 0),
    ("kaynak_notu", "Kaynak Dosyadaki Not / Uyarı", "TEXT", 0, None, 0),
]


def _header(value):
    value = "" if value is None else str(value)
    value = value.replace("I", "ı").replace("İ", "i").lower()
    value = value.translate(str.maketrans("çğıöşü", "cgiosu"))
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9]+", " ", value).strip()


def _spec(key, label, kind, *aliases):
    return {
        "key": key,
        "label": label,
        "kind": kind,
        "aliases": tuple(_header(alias) for alias in aliases),
    }


SOURCE_SPECS = [
    _spec("sira_no", "Sıra No", "int", "Sıra No"),
    _spec(
        "defter_sira_no",
        "Sorumlu Personel",
        "text",
        "Defter Sıra No",
        "Sorumlu Personel",
    ),
    _spec("yil", "Yıl", "int", "Yılı (2017-2021)", "Yıl", "Yılı"),
    _spec("il_adi", "İlin Adı (Sorumlu Teşkilat)", "text", "İlin Adı (Sorumlu Teşkilat)"),
    _spec("hukuki_dayanak", "Hukuki Dayanak", "text", "Hukuki Dayanak"),
    _spec("olcu_aleti_sayisi", "Ölçü Aleti Sayısı", "int", "Ölçü Aleti Sayısı"),
    _spec("olcu_aleti_cinsi", "Ölçü Aleti Cinsi", "text", "Ölçü Aleti Cinsi"),
    _spec("tablet_tutanak_no", "Tablet Tutanak Numarası", "text", "Tablet Tutanak Numarası"),
    _spec(
        "cezanin_muhatabi",
        "Cezanın Muhatabı (Kişi/Firma Adı)",
        "text",
        "Cezanın Muhatabı (Kişi/ Firma Adı)",
        "Cezanın Muhatabı (Kişi/Firma Adı)",
    ),
    _spec(
        "ceza_onay_tarihi",
        "Cezanın Onay Tarihi",
        "date",
        "Cezanın Onay Tarihi (Kararın Verildiği Tarih, İYK Tarihi)",
        "Cezanın Onay Tarihi (İYK Tarihi)",
    ),
    _spec("ceza_tutari", "Verilen Cezanın Tutarı (TL)", "money", "Verilen Cezanın Tutarı (TL)"),
    _spec("gonderim_turu", "Gönderim Türü", "text", "Gönderim Türü"),
    _spec(
        "teblig_tarihi",
        "Cezanın İlgiliye Tebliğ Tarihi",
        "date",
        "Cezanın İlgiliye Tebliğ Tarihi",
    ),
    _spec(
        "dogrudan_yatirilan",
        "Tebligat Üzerine Doğrudan Yatırılan Tutar",
        "money",
        "Tebligat Üzerine İlgili Tarafından Doğrudan Yatırılan Tutar (TL)",
    ),
    _spec("odeme_tarihi", "Ödeme Tarihi", "date", "Ödeme Tarihi"),
    _spec(
        "kesinlesme_tarihi",
        "İtiraz Edilmeyenlerde Kesinleşme Tarihi",
        "date",
        "İtiraz Edilmeyenlerde Kesinleşme Tarihi",
    ),
    _spec("kalan_sure", "Kalan Süre (sistemde yeniden hesaplanır)", "text", "Kalan Süre"),
    _spec(
        "vergi_bildirim_tarihi",
        "Vergi Dairesine Bildirim Tarihi",
        "date",
        "İtiraz Edilmeyenlerde Tahsilat İçin İlgili Vergi Dairesine Bildirim Tarihi",
    ),
    _spec(
        "vergi_dairesi",
        "Bildirim Yapılan Vergi Dairesinin Adı",
        "text",
        "Bildirim Yapılan Vergi Dairesinin Adı",
    ),
    _spec(
        "itiraz_sonucu",
        "İtirazın Sonucu",
        "text",
        "İtirazın Sonucu (Kurum Lehine/ Kurum Aleyhine)",
        "İtirazın Sonucu (Kurum Lehine / Kurum Aleyhine)",
    ),
    _spec(
        "itiraz_lehine_tutar",
        "İtiraz Sonrası Doğrudan Yatırılan Tutar",
        "money",
        "İtiraz Üzerine Kurum Lehine Sonuçlananlarda İlgili Tarafından Doğrudan Yatırılan Tutar (TL)",
        "İtiraz Üzerine Kurum Lehine Sonuçlananlarda Doğrudan Yatırılan Tutar (TL)",
    ),
    _spec(
        "itiraz_vergi_bildirim_tarihi",
        "İtiraz Sonrası Vergi Dairesine Bildirim Tarihi",
        "date",
        "İtiraz Üzerine Kurum Lehine Sonuçlananlarda Tahsilat İçin İlgili Vergi Dairesine Bildirim Tarihi",
        "İtiraz Üzerine Kurum Lehine Sonuçlananlarda Vergi Dairesine Bildirim Tarihi",
    ),
    _spec(
        "vergi_tahsil_tutari",
        "Vergi Dairesince Tahsil Edildiği Bildirilen Tutar",
        "money",
        "İlgili Vergi Dairesince Tahsil Edildiği Bildirilen Tutar (TL)",
    ),
    _spec(
        "vergi_tahsil_tarihi",
        "Vergi Dairesince Tahsil Edildiği Bildirilen Tarih",
        "date",
        "İlgili Vergi Dairesince Tahsil Edildiği Bildirilen Tarih",
    ),
    _spec("tespit_kurum", "Tespiti Yapan Kurum", "text", "Tespiti Yapan Kurum", "Tespit Yapan Kurum"),
    _spec(
        "ebys",
        "Tespit Yazısının EBYS Tarih ve Sayısı",
        "text",
        "Tespit Yazısının EBYS Tarih ve Sayısı",
    ),
    _spec("kaynak_durumu", "Kaynak Dosyadaki Durum", "text", "Durumu"),
]

SPEC_BY_KEY = {spec["key"]: spec for spec in SOURCE_SPECS}
OPTION_KEYS = {
    "hukuki_dayanak",
    "gonderim_turu",
    "vergi_dairesi",
    "tespit_kurum",
}
# Some source workbooks pre-fill future rows with a sequence number, year,
# institution, legal basis and formulas. Those defaults do not make a real IPC
# record. At least one business-specific value must exist before a row can be
# imported.
RECORD_EVIDENCE_KEYS = {
    "defter_sira_no",
    "olcu_aleti_sayisi",
    "olcu_aleti_cinsi",
    "tablet_tutanak_no",
    "cezanin_muhatabi",
    "ceza_onay_tarihi",
    "ceza_tutari",
    "gonderim_turu",
    "teblig_tarihi",
    "dogrudan_yatirilan",
    "odeme_tarihi",
    "vergi_bildirim_tarihi",
    "vergi_dairesi",
    "itiraz_sonucu",
    "itiraz_lehine_tutar",
    "itiraz_vergi_bildirim_tarihi",
    "vergi_tahsil_tutari",
    "vergi_tahsil_tarihi",
    "tespit_kurum",
    "ebys",
    "kaynak_durumu",
}
CORE_KEYS = {"sira_no", "yil", "hukuki_dayanak", "cezanin_muhatabi", "ceza_tutari"}
MAX_COLUMNS = 200
MAX_ROWS = 100_000
BLANK_ROW_STOP = 25
WARNING_SAMPLE_LIMIT = 200


def normalize_instrument_items(items=None, quantity="", kind=""):
    """Return ordered, merged instrument type/positive-quantity pairs."""
    candidates = []
    if isinstance(items, list):
        candidates = items
    else:
        kind_text = _clean_text(kind)
        parts = [part.strip() for part in kind_text.split(";") if part.strip()]
        parsed_parts = []
        if parts:
            for part in parts:
                match = re.fullmatch(r"(.+?)\s*\((\d+)\)", part)
                if not match or int(match.group(2)) <= 0:
                    parsed_parts = []
                    break
                parsed_parts.append(
                    {"cinsi": match.group(1).strip(), "sayisi": int(match.group(2))}
                )
        if parsed_parts:
            candidates = parsed_parts
        else:
            cleaned_quantity, quantity_error = _clean_int(quantity)
            if kind_text and not quantity_error and cleaned_quantity:
                candidates = [
                    {"cinsi": kind_text, "sayisi": int(cleaned_quantity)}
                ]

    merged = {}
    order = []
    for item in candidates:
        if not isinstance(item, dict):
            continue
        item_kind = _clean_text(item.get("cinsi"))
        try:
            item_quantity = int(str(item.get("sayisi", "")).strip())
        except (TypeError, ValueError):
            continue
        if not item_kind or item_quantity <= 0:
            continue
        if item_kind not in merged:
            merged[item_kind] = 0
            order.append(item_kind)
        merged[item_kind] += item_quantity
    return [
        {"cinsi": item_kind, "sayisi": merged[item_kind]}
        for item_kind in order
    ]


def instrument_total(items):
    return str(sum(item["sayisi"] for item in items)) if items else ""


def instrument_summary(items):
    return "; ".join(
        f"{item['cinsi']} ({item['sayisi']})" for item in items
    )


def _compact_row_ranges(row_numbers):
    """Return human-readable Excel row ranges without losing gaps."""
    if not row_numbers:
        return ""
    ordered = sorted(set(row_numbers))
    ranges = []
    start = previous = ordered[0]
    for row_number in ordered[1:]:
        if row_number == previous + 1:
            previous = row_number
            continue
        ranges.append(str(start) if start == previous else f"{start}–{previous}")
        start = previous = row_number
    ranges.append(str(start) if start == previous else f"{start}–{previous}")
    return ", ".join(ranges)


def file_sha256(path):
    digest = hashlib.sha256()
    with open(path, "rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _match_spec(header_value):
    normalized = _header(header_value)
    if not normalized:
        return None
    for spec in SOURCE_SPECS:
        for alias in spec["aliases"]:
            if normalized == alias:
                return spec
            if len(alias) >= 18 and (alias in normalized or normalized in alias):
                return spec
    return None


def _clean_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    return re.sub(r"\s+", " ", str(value)).strip()


def _clean_int(value):
    if value is None or value == "":
        return "", None
    if isinstance(value, bool):
        return "", "tam sayı değil"
    if isinstance(value, (int, float, Decimal)):
        number = Decimal(str(value))
        if number == number.to_integral_value():
            return str(int(number)), None
        return "", "tam sayı değil"
    text = _clean_text(value).replace(" ", "")
    if re.fullmatch(r"[+-]?\d+(?:[.,]0+)?", text):
        return str(int(Decimal(text.replace(",", ".")))), None
    return "", "tam sayı değil"


def _clean_money(value):
    if value is None or value == "":
        return "", None, None
    if isinstance(value, bool):
        return "", None, "tutar değil"
    changed = None
    if isinstance(value, (int, float, Decimal)):
        number = Decimal(str(value))
    else:
        original = _clean_text(value)
        text = original.upper().replace("₺", "").replace("TL", "")
        text = text.replace("\u00a0", "").replace(" ", "")
        text = re.sub(r"[^0-9,.\-+]", "", text)
        if not text or not re.search(r"\d", text):
            return "", None, "tutar değil"
        if "," in text and "." in text:
            if text.rfind(",") > text.rfind("."):
                text = text.replace(".", "").replace(",", ".")
            else:
                text = text.replace(",", "")
        elif "," in text:
            text = text.replace(".", "").replace(",", ".")
        elif text.count(".") > 1:
            text = text.replace(".", "")
        try:
            number = Decimal(text)
        except InvalidOperation:
            return "", None, "tutar değil"
        if original != text:
            changed = "Metin olarak yazılmış tutarlar sayıya dönüştürüldü"
    normalized = format(number, "f")
    if "." in normalized:
        normalized = normalized.rstrip("0").rstrip(".")
    if normalized in ("", "-0"):
        normalized = "0"
    return normalized, changed, None


def _clean_date(value):
    if value is None or value == "":
        return "", None
    if isinstance(value, datetime):
        return value.date().strftime("%d.%m.%Y"), None
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y"), None
    if isinstance(value, (int, float, Decimal)) and 1 <= float(value) <= 100_000:
        try:
            converted = from_excel(float(value))
            return converted.strftime("%d.%m.%Y"), None
        except (TypeError, ValueError, OverflowError):
            pass
    text = _clean_text(value)
    text = re.sub(r"\s+00:00:00$", "", text)
    formats = (
        "%d.%m.%Y",
        "%d.%m.%y",
        "%m/%d/%Y",
        "%m/%d/%y",
        "%d/%m/%Y",
        "%d/%m/%y",
        "%Y-%m-%d",
    )
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).strftime("%d.%m.%Y"), None
        except ValueError:
            pass
    return "", "geçerli tarih değil"


def _normalize_value(value, kind):
    if kind == "text":
        return _clean_text(value), None, None
    if kind == "int":
        cleaned, error = _clean_int(value)
        return cleaned, None, error
    if kind == "money":
        return _clean_money(value)
    if kind == "date":
        cleaned, error = _clean_date(value)
        return cleaned, None, error
    return _clean_text(value), None, None


def _find_sheet_and_header(workbook):
    best = None
    for sheet in workbook.worksheets:
        max_col = min(max(sheet.max_column or 1, 30), MAX_COLUMNS)
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=1, max_row=20, max_col=max_col, values_only=True),
            start=1,
        ):
            matched = {_match_spec(value)["key"] for value in row if _match_spec(value)}
            candidate = (len(matched), sheet, row_number, row)
            if best is None or candidate[0] > best[0]:
                best = candidate
    if not best or best[0] < len(CORE_KEYS):
        raise ImportWorkbookError(
            "Dosyada beklenen İPC başlıkları bulunamadı. "
            "En az Sıra No, Yıl, Hukuki Dayanak, Cezanın Muhatabı ve Ceza Tutarı sütunları gerekir."
        )
    _, sheet, row_number, row = best
    found_core = {_match_spec(value)["key"] for value in row if _match_spec(value)}
    missing = CORE_KEYS - found_core
    if missing:
        labels = ", ".join(SPEC_BY_KEY[key]["label"] for key in sorted(missing))
        raise ImportWorkbookError(f"Zorunlu kaynak sütunları eksik: {labels}.")
    return sheet, row_number, row


def parse_workbook(path):
    """Parse an İPC workbook without changing it or the application database."""
    path = Path(path)
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ImportWorkbookError(f"Excel dosyası açılamadı: {exc}") from exc

    warning_samples = []
    warning_count = 0
    normalization_counts = Counter()

    def warn(row_number, column, message):
        nonlocal warning_count
        warning_count += 1
        if len(warning_samples) < WARNING_SAMPLE_LIMIT:
            warning_samples.append(
                {"row": row_number, "column": column, "message": message}
            )

    try:
        sheet, header_row, header_values = _find_sheet_and_header(workbook)
        max_col = min(max(sheet.max_column or 1, len(header_values), 30), MAX_COLUMNS)
        header_values = tuple(header_values) + (None,) * (max_col - len(header_values))

        mapping = {}
        mapped_keys = set()
        mapping_report = []
        for index, raw_header in enumerate(header_values[:max_col]):
            spec = _match_spec(raw_header)
            if spec and spec["key"] not in mapped_keys:
                mapping[index] = spec
                mapped_keys.add(spec["key"])
                mapping_report.append(
                    {
                        "column": get_column_letter(index + 1),
                        "source": _clean_text(raw_header),
                        "target": spec["label"],
                    }
                )

        records = []
        option_values = defaultdict(set)
        years = Counter()
        total_amount = Decimal("0")
        nonempty_row_numbers = []
        template_row_numbers = []
        rejected_row_numbers = []
        blank_streak = 0
        started = False

        rows = sheet.iter_rows(
            min_row=header_row + 1,
            max_row=MAX_ROWS,
            max_col=max_col,
            values_only=True,
        )
        for row_number, row in enumerate(rows, start=header_row + 1):
            if not any(value not in (None, "") for value in row):
                if started:
                    blank_streak += 1
                    if blank_streak >= BLANK_ROW_STOP:
                        break
                continue
            started = True
            blank_streak = 0
            nonempty_row_numbers.append(row_number)

            sira_col = next(
                (index for index, spec in mapping.items() if spec["key"] == "sira_no"),
                None,
            )
            source_sira, sira_error = _clean_int(row[sira_col] if sira_col is not None else None)
            if sira_error or not source_sira:
                rejected_row_numbers.append(row_number)
                warn(row_number, get_column_letter((sira_col or 0) + 1), "Sıra No okunamadığı için satır aktarılmadı.")
                continue

            data = {}
            for index, spec in mapping.items():
                raw_value = row[index] if index < len(row) else None
                if spec["key"] == "sira_no":
                    continue
                cleaned, normalization, error = _normalize_value(raw_value, spec["kind"])
                if error and raw_value not in (None, ""):
                    warn(
                        row_number,
                        get_column_letter(index + 1),
                        f"{spec['label']}: “{_clean_text(raw_value)}” {error}; ham değer not alanında korundu.",
                    )
                    data.setdefault("kaynak_notu", "")
                    note = f"{get_column_letter(index + 1)} ({spec['label']}): {_clean_text(raw_value)}"
                    data["kaynak_notu"] = "; ".join(filter(None, [data["kaynak_notu"], note]))
                    cleaned = ""
                if normalization:
                    normalization_counts[normalization] += 1
                data[spec["key"]] = cleaned
                if spec["key"] in OPTION_KEYS and cleaned:
                    option_values[spec["key"]].add(cleaned)

            instrument_items = normalize_instrument_items(
                quantity=data.get("olcu_aleti_sayisi", ""),
                kind=data.get("olcu_aleti_cinsi", ""),
            )
            if instrument_items:
                source_total = data.get("olcu_aleti_sayisi", "")
                calculated_total = instrument_total(instrument_items)
                if source_total and source_total != calculated_total:
                    warn(
                        row_number,
                        "F",
                        "Ölçü Aleti Sayısı, G sütunundaki kalemlerin toplamına göre düzeltildi.",
                    )
                data["olcu_aletleri"] = instrument_items
                data["olcu_aleti_sayisi"] = calculated_total
                data["olcu_aleti_cinsi"] = instrument_summary(instrument_items)
                for item in instrument_items:
                    option_values["olcu_aleti_cinsi"].add(item["cinsi"])
            elif data.get("olcu_aleti_cinsi"):
                option_values["olcu_aleti_cinsi"].add(data["olcu_aleti_cinsi"])

            extra_notes = []
            for index, raw_value in enumerate(row):
                if index in mapping or raw_value in (None, ""):
                    continue
                column = get_column_letter(index + 1)
                header = _clean_text(header_values[index]) if index < len(header_values) else ""
                label = f"{column} ({header})" if header else column
                extra_notes.append(f"{label}: {_clean_text(raw_value)}")
            if extra_notes:
                normalization_counts["Başlıksız / eşleşmeyen hücreler kaynak notunda korundu"] += len(extra_notes)
                data["kaynak_notu"] = "; ".join(
                    filter(None, [data.get("kaynak_notu", ""), *extra_notes])
                )

            if not any(data.get(key) not in (None, "") for key in RECORD_EVIDENCE_KEYS):
                template_row_numbers.append(row_number)
                normalization_counts["Boş şablon satırı aktarılmadan atlandı"] += 1
                continue

            source_number = int(source_sira)
            records.append(
                {"sira_no": source_number, "source_row": row_number, "data": data}
            )
            if data.get("yil"):
                years[data["yil"]] += 1
            if data.get("ceza_tutari"):
                try:
                    total_amount += Decimal(data["ceza_tutari"])
                except InvalidOperation:
                    pass

        if not records:
            raise ImportWorkbookError("Başlıklar bulundu ancak aktarılabilecek hiçbir kayıt bulunamadı.")

        duplicate_source_ids = [
            sira for sira, count in Counter(r["sira_no"] for r in records).items() if count > 1
        ]
        if duplicate_source_ids:
            sample = ", ".join(str(value) for value in duplicate_source_ids[:10])
            raise ImportWorkbookError(
                f"Kaynak dosyada yinelenen Sıra No değerleri var ({sample}). "
                "Hangi kaydın doğru olduğu belirlenmeden aktarım yapılamaz."
            )

        mapped_columns = {item["column"] for item in mapping_report}
        unrecognized_headers = []
        for index, raw_header in enumerate(header_values[:max_col]):
            if not raw_header or get_column_letter(index + 1) in mapped_columns:
                continue
            unrecognized_headers.append(
                {"column": get_column_letter(index + 1), "header": _clean_text(raw_header)}
            )

        return {
            "filename": path.name,
            "digest": file_sha256(path),
            "sheet_name": sheet.title,
            "header_row": header_row,
            "records": records,
            "record_count": len(records),
            "source_nonempty_row_count": len(nonempty_row_numbers),
            "ignored_row_count": len(template_row_numbers) + len(rejected_row_numbers),
            "template_row_count": len(template_row_numbers),
            "rejected_row_count": len(rejected_row_numbers),
            "source_row_range": _compact_row_ranges(nonempty_row_numbers),
            "record_row_range": _compact_row_ranges(
                record["source_row"] for record in records
            ),
            "template_row_range": _compact_row_ranges(template_row_numbers),
            "rejected_row_range": _compact_row_ranges(rejected_row_numbers),
            "mapping": mapping_report,
            "mapped_count": len(mapping_report),
            "unrecognized_headers": unrecognized_headers,
            "warning_count": warning_count,
            "warning_samples": warning_samples,
            "normalizations": dict(normalization_counts),
            "years": dict(sorted(years.items())),
            "total_amount": format(total_amount, "f"),
            "option_values": {key: sorted(values) for key, values in option_values.items()},
        }
    finally:
        workbook.close()
