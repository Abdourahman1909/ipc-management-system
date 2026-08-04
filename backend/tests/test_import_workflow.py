# -*- coding: utf-8 -*-

import io
import json
import os
import re
import sqlite3
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from test_importer import HEADERS


class ImportWorkflowTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.temp_dir = tempfile.TemporaryDirectory()
        root = Path(cls.temp_dir.name)
        os.environ["IPC_DB_PATH"] = str(root / "test-ipc.db")
        os.environ["IPC_IMPORT_TMP_DIR"] = str(root / "imports")
        os.environ["IPC_BACKUP_DIR"] = str(root / "backups")
        os.environ["IPC_UPLOAD_DIR"] = str(root / "uploads")
        os.environ["IPC_INITIAL_ADMIN_EMAIL"] = "admin@example.local"
        os.environ["IPC_INITIAL_ADMIN_PASSWORD"] = "test-only-password"

        import app as app_module

        cls.module = app_module
        cls.app = app_module.app
        cls.app.config.update(TESTING=True)

    @classmethod
    def tearDownClass(cls):
        cls.temp_dir.cleanup()

    def setUp(self):
        database = sqlite3.connect(self.module.DB_PATH)
        database.execute("DELETE FROM attachments")
        database.execute("DELETE FROM pending_actions")
        database.execute("DELETE FROM notifications")
        database.execute("DELETE FROM companies")
        database.execute("DELETE FROM import_history")
        database.execute("DELETE FROM users WHERE id<>1")
        database.commit()
        database.close()
        upload_dir = Path(self.module.UPLOAD_DIR)
        if upload_dir.exists():
            for path in upload_dir.iterdir():
                if path.is_file():
                    path.unlink()
        self.client = self.app.test_client()
        with self.client.session_transaction() as session:
            session["user_id"] = 1
            session["user_name"] = "admin"
            session["role"] = "admin"

    def workbook_bytes(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "İPC Bilgi Girişi"
        sheet.append(["İDARİ PARA CEZALARINA İLİŞKİN TABLO"])
        sheet.append(HEADERS)
        for number, amount in ((1, 1000), (2, 2500)):
            row = [None] * 30
            row[0] = number
            row[2] = 2026
            row[3] = "KOCAELİ"
            row[4] = "3516.15/c"
            row[6] = "Taksimetre"
            row[9] = f"Örnek Firma {number}"
            row[11] = amount
            row[12] = "UETS"
            row[19] = "İLYASBEY VD."
            row[25] = "TSE"
            sheet.append(row)
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer

    def preview(self):
        response = self.client.post(
            "/veri-aktar",
            data={"workbook": (self.workbook_bytes(), "tarihsel-ipc.xlsx")},
            content_type="multipart/form-data",
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("Önizleme".encode(), response.data)
        self.assertIn("Gerçek İPC kaydı".encode(), response.data)
        self.assertIn("Dolu görünen satır".encode(), response.data)
        self.assertIn("Atlanan boş şablon".encode(), response.data)
        with self.client.session_transaction() as session:
            return session["ipc_import"]["token"]

    def valid_record_form(self):
        today = self.module.date.today().strftime("%d.%m.%Y")
        return {
            "yil": str(self.module.date.today().year),
            "il_adi": "KOCAELİ",
            "hukuki_dayanak": "3516.15/c",
            "olcu_aleti_sayisi": "1",
            "olcu_aleti_cinsi": "Taksimetre",
            "tablet_tutanak_no": "41-E7E8EC31",
            "cezanin_muhatabi": "Ek Dosya Test Firması",
            "ceza_onay_tarihi": today,
            "ceza_tutari": "1000",
            "gonderim_turu": "UETS",
            "teblig_tarihi": today,
            "tespit_kurum": "TSE",
            "ebys": f"{today}-1234567",
        }

    def test_preview_does_not_write_and_confirm_imports_with_backup(self):
        token = self.preview()
        database = sqlite3.connect(self.module.DB_PATH)
        self.assertEqual(database.execute("SELECT COUNT(*) FROM companies").fetchone()[0], 0)
        database.close()

        response = self.client.post(
            "/veri-aktar/onayla",
            data={"token": token, "conflict_mode": "skip"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("Aktarım tamamlandı".encode(), response.data)

        database = sqlite3.connect(self.module.DB_PATH)
        database.row_factory = sqlite3.Row
        rows = database.execute("SELECT * FROM companies ORDER BY sira_no").fetchall()
        self.assertEqual(len(rows), 2)
        self.assertEqual(json.loads(rows[0]["data"])["ceza_tutari"], "1000")
        self.assertEqual(database.execute("SELECT COUNT(*) FROM import_history").fetchone()[0], 1)
        field = database.execute(
            "SELECT label FROM form_fields WHERE field_key='kaynak_durumu'"
        ).fetchone()
        self.assertIsNotNone(field)
        database.close()
        self.assertTrue(any(Path(self.module.BACKUP_DIR).glob("ipc-before-import-*.db")))

    def test_skip_mode_keeps_existing_conflict(self):
        database = sqlite3.connect(self.module.DB_PATH)
        database.execute(
            "INSERT INTO companies(sira_no,data,created_by) VALUES(1,?,?)",
            (json.dumps({"ceza_tutari": "99", "personel_adi": "Mevcut"}), "test"),
        )
        database.commit()
        database.close()

        token = self.preview()
        response = self.client.post(
            "/veri-aktar/onayla",
            data={"token": token, "conflict_mode": "skip"},
        )
        self.assertEqual(response.status_code, 200)

        database = sqlite3.connect(self.module.DB_PATH)
        value = database.execute(
            "SELECT data FROM companies WHERE sira_no=1"
        ).fetchone()[0]
        self.assertEqual(json.loads(value)["ceza_tutari"], "99")
        self.assertEqual(database.execute("SELECT COUNT(*) FROM companies").fetchone()[0], 2)
        database.close()

    def test_deadline_clock_stops_for_closed_historical_records(self):
        today = self.module.date.today().strftime("%d.%m.%Y")
        open_record = {"teblig_tarihi": today}
        paid_record = {
            "teblig_tarihi": "01.01.2017",
            "odeme_tarihi": "10.01.2017",
        }
        source_paid_record = {
            "teblig_tarihi": "01.01.2017",
            "kaynak_durumu": "ÖDEMESİ YAPILDI",
        }
        source_unpaid_record = {
            "teblig_tarihi": "01.01.2017",
            "kaynak_durumu": "ÖDEMESİ YAPILMADI",
        }

        self.assertEqual(self.module.kalan_sure(open_record), 30)
        self.assertIsNone(self.module.kalan_sure(paid_record))
        self.assertIsNone(self.module.kalan_sure(source_paid_record))
        self.assertIsInstance(self.module.kalan_sure(source_unpaid_record), int)

    def test_sqlite_is_configured_for_small_lan_concurrency(self):
        with self.app.app_context():
            database = self.module.get_db()
            self.assertEqual(
                database.execute("PRAGMA journal_mode").fetchone()[0],
                "wal",
            )
            self.assertEqual(
                database.execute("PRAGMA busy_timeout").fetchone()[0],
                30000,
            )

    def test_registry_searches_tablet_number_and_company_name(self):
        database = sqlite3.connect(self.module.DB_PATH)
        records = [
            (
                10,
                {
                    "yil": "2026",
                    "tablet_tutanak_no": "41-ABCDEF12",
                    "cezanin_muhatabi": "Körfez Ölçüm Ltd.",
                    "hukuki_dayanak": "3516.15/c",
                    "ceza_tutari": "16000",
                },
            ),
            (
                11,
                {
                    "yil": "2025",
                    "tablet_tutanak_no": "41-87654321",
                    "cezanin_muhatabi": "Gebze Sanayi A.Ş.",
                    "hukuki_dayanak": "3516.15/e",
                    "ceza_tutari": "25000",
                },
            ),
        ]
        database.executemany(
            "INSERT INTO companies(sira_no,data,created_by) VALUES(?,?,?)",
            [
                (number, json.dumps(data, ensure_ascii=False), "test")
                for number, data in records
            ],
        )
        database.commit()
        database.close()

        by_tablet = self.client.get("/sirketler?q=abcdef")
        self.assertEqual(by_tablet.status_code, 200)
        self.assertIn("Körfez Ölçüm".encode(), by_tablet.data)
        self.assertNotIn("Gebze Sanayi".encode(), by_tablet.data)

        by_name = self.client.get("/sirketler?q=GEBZE")
        self.assertEqual(by_name.status_code, 200)
        self.assertIn("Gebze Sanayi".encode(), by_name.data)
        self.assertNotIn("Körfez Ölçüm".encode(), by_name.data)

    def test_unified_record_workspace_opens_view_and_edit_pages(self):
        database = sqlite3.connect(self.module.DB_PATH)
        database.execute(
            "INSERT INTO companies(sira_no,data,created_by) VALUES(?,?,?)",
            (
                12,
                json.dumps(
                    {
                        "yil": "2026",
                        "tablet_tutanak_no": "41-A1B2C3D4",
                        "cezanin_muhatabi": "Birleşik Menü Testi",
                        "hukuki_dayanak": "3516.15/c",
                        "ceza_tutari": "12000",
                    },
                    ensure_ascii=False,
                ),
                "admin",
            ),
        )
        database.commit()
        database.close()

        listing = self.client.get("/sirketler")
        self.assertEqual(listing.status_code, 200)
        self.assertIn("İPC Kayıtları".encode(), listing.data)
        self.assertNotIn("Şirket Bilgisi Sil".encode(), listing.data)
        self.assertNotIn("Şirket Bilgisi Düzenle".encode(), listing.data)
        self.assertNotIn("Şirket Bilgisi Oku".encode(), listing.data)

        detail = self.client.get("/sirket-oku?sira_no=12")
        self.assertEqual(detail.status_code, 200)
        self.assertIn("Birleşik Menü Testi".encode(), detail.data)
        self.assertIn("Kaydı düzenle".encode(), detail.data)
        self.assertIn("Kaydı sil".encode(), detail.data)

        edit = self.client.get("/sirket-duzenle?sira_no=12")
        self.assertEqual(edit.status_code, 200)
        self.assertIn("İPC Kaydını Düzenle".encode(), edit.data)
        self.assertIn("Tebliğ tarihinden itibaren 30 gün".encode(), edit.data)
        self.assertIn("Bir veya birden fazla belge".encode(), edit.data)

    def test_report_builder_previews_filtered_records(self):
        database = sqlite3.connect(self.module.DB_PATH)
        database.executemany(
            "INSERT INTO companies(sira_no,data,created_by) VALUES(?,?,?)",
            [
                (
                    20,
                    json.dumps(
                        {
                            "yil": "2026",
                            "tablet_tutanak_no": "41-AAAABBBB",
                            "cezanin_muhatabi": "Rapor Hedef Firma",
                            "ceza_tutari": "1000",
                        },
                        ensure_ascii=False,
                    ),
                    "admin",
                ),
                (
                    21,
                    json.dumps(
                        {
                            "yil": "2025",
                            "tablet_tutanak_no": "41-CCCCDDDD",
                            "cezanin_muhatabi": "Rapor Dışı Firma",
                            "ceza_tutari": "2000",
                        },
                        ensure_ascii=False,
                    ),
                    "admin",
                ),
            ],
        )
        database.commit()
        database.close()

        response = self.client.get(
            "/rapor?rapor_turu=ozet&q=hedef&onizle=1"
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn("Yönetim Özeti".encode(), response.data)
        self.assertIn("Rapor Hedef Firma".encode(), response.data)
        self.assertNotIn("Rapor Dışı Firma".encode(), response.data)
        self.assertIn("1.000,00 TL".encode(), response.data)

    def test_report_builder_filters_an_inclusive_year_range(self):
        database = sqlite3.connect(self.module.DB_PATH)
        database.executemany(
            "INSERT INTO companies(sira_no,data,created_by) VALUES(?,?,?)",
            [
                (
                    24,
                    json.dumps(
                        {
                            "yil": "2024",
                            "cezanin_muhatabi": "Aralık Öncesi Firma",
                            "ceza_tutari": "1000",
                        },
                        ensure_ascii=False,
                    ),
                    "admin",
                ),
                (
                    25,
                    json.dumps(
                        {
                            "yil": "2025",
                            "cezanin_muhatabi": "Aralık İçinde Birinci Firma",
                            "ceza_tutari": "2000",
                        },
                        ensure_ascii=False,
                    ),
                    "admin",
                ),
                (
                    26,
                    json.dumps(
                        {
                            "yil": "2026",
                            "cezanin_muhatabi": "Aralık İçinde İkinci Firma",
                            "ceza_tutari": "3000",
                        },
                        ensure_ascii=False,
                    ),
                    "admin",
                ),
            ],
        )
        database.commit()
        database.close()

        response = self.client.get(
            "/rapor?rapor_turu=ozet&yil_baslangic=2025&yil_bitis=2026&onizle=1"
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn("Aralık İçinde Birinci Firma".encode(), response.data)
        self.assertIn("Aralık İçinde İkinci Firma".encode(), response.data)
        self.assertNotIn("Aralık Öncesi Firma".encode(), response.data)
        self.assertIn("5.000,00 TL".encode(), response.data)

    def test_report_preview_orders_records_chronologically(self):
        database = sqlite3.connect(self.module.DB_PATH)
        database.executemany(
            "INSERT INTO companies(sira_no,data,created_by) VALUES(?,?,?)",
            [
                (
                    30,
                    json.dumps(
                        {"yil": "2026", "cezanin_muhatabi": "Son Kayıt"},
                        ensure_ascii=False,
                    ),
                    "admin",
                ),
                (
                    2,
                    json.dumps(
                        {"yil": "2017", "cezanin_muhatabi": "İkinci Kayıt"},
                        ensure_ascii=False,
                    ),
                    "admin",
                ),
                (
                    1,
                    json.dumps(
                        {"yil": "2017", "cezanin_muhatabi": "İlk Kayıt"},
                        ensure_ascii=False,
                    ),
                    "admin",
                ),
            ],
        )
        database.commit()
        database.close()

        response = self.client.get("/rapor?rapor_turu=tam&onizle=1")

        self.assertEqual(response.status_code, 200)
        self.assertLess(
            response.data.index("İlk Kayıt".encode()),
            response.data.index("İkinci Kayıt".encode()),
        )
        self.assertLess(
            response.data.index("İkinci Kayıt".encode()),
            response.data.index("Son Kayıt".encode()),
        )

    def test_full_pdf_uses_readable_record_pages_instead_of_28_column_table(self):
        database = sqlite3.connect(self.module.DB_PATH)
        database.executemany(
            "INSERT INTO companies(sira_no,data,created_by) VALUES(?,?,?)",
            [
                (
                    number,
                    json.dumps(
                        {
                            "yil": "2026",
                            "il_adi": "KOCAELİ",
                            "hukuki_dayanak": "3516.15/c",
                            "tablet_tutanak_no": f"41-TEST{number:04d}",
                            "cezanin_muhatabi": f"PDF Test Firması {number}",
                            "ceza_tutari": "1000",
                            "teblig_tarihi": "01.07.2026",
                        },
                        ensure_ascii=False,
                    ),
                    "admin",
                )
                for number in (41, 42)
            ],
        )
        database.commit()
        database.close()

        full_pdf = self.client.get("/rapor/pdf?rapor_turu=tam")
        compact_pdf = self.client.get("/rapor/pdf?rapor_turu=ozet")

        self.assertEqual(full_pdf.status_code, 200)
        self.assertEqual(compact_pdf.status_code, 200)
        self.assertTrue(full_pdf.data.startswith(b"%PDF"))
        self.assertTrue(compact_pdf.data.startswith(b"%PDF"))
        self.assertEqual(
            len(re.findall(rb"/Type /Page\b", full_pdf.data)),
            2,
        )
        self.assertEqual(
            len(re.findall(rb"/Type /Page\b", compact_pdf.data)),
            1,
        )

    def test_notifications_group_actions_and_do_not_mark_all_read_on_view(self):
        database = sqlite3.connect(self.module.DB_PATH)
        database.executemany(
            "INSERT INTO companies(sira_no,data,created_by) VALUES(?,?,?)",
            [
                (
                    80,
                    json.dumps(
                        {
                            "cezanin_muhatabi": "Süre Testi Bir",
                            "teblig_tarihi": "01.01.2017",
                        },
                        ensure_ascii=False,
                    ),
                    "admin",
                ),
                (
                    81,
                    json.dumps(
                        {
                            "cezanin_muhatabi": "Süre Testi İki",
                            "teblig_tarihi": "02.01.2017",
                        },
                        ensure_ascii=False,
                    ),
                    "admin",
                ),
            ],
        )
        database.execute(
            """INSERT INTO notifications(user_id,title,message)
               VALUES(1,'Bilgilendirme','Tekil bilgi mesajı')"""
        )
        database.commit()
        database.close()

        response = self.client.get("/bildirimler")

        self.assertEqual(response.status_code, 200)
        self.assertIn("2 kayıt için süre dolmuş görünüyor.".encode(), response.data)
        self.assertIn("Son gelişmeler".encode(), response.data)

        database = sqlite3.connect(self.module.DB_PATH)
        grouped = database.execute(
            """SELECT COUNT(*),is_read FROM notifications
               WHERE dedupe_key='deadline-expired'"""
        ).fetchone()
        info_read = database.execute(
            "SELECT is_read FROM notifications WHERE message='Tekil bilgi mesajı'"
        ).fetchone()[0]
        self.assertEqual(grouped, (1, 0))
        self.assertEqual(info_read, 0)
        database.close()

        second_response = self.client.get("/bildirimler")
        self.assertEqual(second_response.status_code, 200)
        database = sqlite3.connect(self.module.DB_PATH)
        self.assertEqual(
            database.execute(
                """SELECT COUNT(*) FROM notifications
                   WHERE dedupe_key='deadline-expired'"""
            ).fetchone()[0],
            1,
        )
        database.close()

        mark_read = self.client.post("/bildirimler/tumunu-okundu")
        self.assertEqual(mark_read.status_code, 302)
        database = sqlite3.connect(self.module.DB_PATH)
        self.assertEqual(
            database.execute(
                """SELECT COUNT(*) FROM notifications
                   WHERE user_id=1 AND is_read=0 AND resolved_at IS NULL"""
            ).fetchone()[0],
            0,
        )
        database.close()

    def test_approval_workspace_shows_only_changes_and_collapses_history(self):
        database = sqlite3.connect(self.module.DB_PATH)
        cursor = database.execute(
            """INSERT INTO users(name,email,password_hash,role)
               VALUES('personel','personel@example.test','x','personel')"""
        )
        personnel_id = cursor.lastrowid
        database.execute(
            "INSERT INTO companies(sira_no,data,created_by) VALUES(?,?,?)",
            (
                15,
                json.dumps(
                    {
                        "cezanin_muhatabi": "Eski Firma",
                        "ceza_tutari": "1000",
                        "yil": "2025",
                    },
                    ensure_ascii=False,
                ),
                "admin",
            ),
        )
        database.execute(
            """INSERT INTO pending_actions
               (action_type,sira_no,data,requested_by,status)
               VALUES('duzenle',15,?,?,'bekliyor')""",
            (
                json.dumps(
                    {
                        "cezanin_muhatabi": "Yeni Firma",
                        "ceza_tutari": "2000",
                        "yil": "2025",
                    },
                    ensure_ascii=False,
                ),
                personnel_id,
            ),
        )
        database.execute(
            """INSERT INTO pending_actions
               (action_type,sira_no,requested_by,status,decision_reason)
               VALUES('sil',15,?,'reddedildi','Kayıt korunmalı.')""",
            (personnel_id,),
        )
        database.commit()
        database.close()

        response = self.client.get("/onaylar")

        self.assertEqual(response.status_code, 200)
        self.assertIn("Onay Talepleri".encode(), response.data)
        self.assertIn("Karar bekleyenler".encode(), response.data)
        self.assertIn("Eski Firma".encode(), response.data)
        self.assertIn("Yeni Firma".encode(), response.data)
        self.assertIn("Geçmiş kararlar".encode(), response.data)
        self.assertIn("Kayıt korunmalı.".encode(), response.data)
        self.assertNotIn(b'title="Onay Bekleyenler"', response.data)

    def test_dashboard_uses_empty_state_instead_of_blank_charts(self):
        response = self.client.get("/panel")
        self.assertEqual(response.status_code, 200)
        self.assertIn("Henüz İPC kaydı bulunmuyor".encode(), response.data)
        self.assertNotIn(b'id="yearTrendChart"', response.data)
        self.assertNotIn(b'id="statusDistributionChart"', response.data)
        self.assertNotIn(b"dashboard-actions", response.data)
        self.assertIn(b"ipc-denetim-terazi.webp", response.data)
        self.assertIn(b"sidebar-brand", response.data)
        self.assertIn(b'img/ministry-emblem.png', response.data)
        self.assertIn("İdari Para Cezaları".encode(), response.data)
        self.assertIn(b"navbar-brand d-flex d-lg-none", response.data)
        self.assertIn(b"favicon-32x32.png", response.data)
        self.assertIn(b'name="theme-color" content="#5C0F1D"', response.data)

    def test_dashboard_visualizes_all_current_records(self):
        today = self.module.date.today()
        database = sqlite3.connect(self.module.DB_PATH)
        records = [
            {
                "yil": str(today.year),
                "ceza_tutari": "1000",
                "dogrudan_yatirilan": "250",
                "teblig_tarihi": today.strftime("%d.%m.%Y"),
            },
            {
                "yil": str(today.year - 1),
                "ceza_tutari": "2000",
                "teblig_tarihi": "01.01.2017",
            },
            {
                "yil": str(today.year - 1),
                "ceza_tutari": "3000",
                "odeme_tarihi": "10.01.2017",
                "vergi_tahsil_tutari": "500",
            },
            {
                "yil": str(today.year),
                "ceza_tutari": "4000",
            },
        ]
        database.executemany(
            "INSERT INTO companies(sira_no,data,created_by) VALUES(?,?,?)",
            [
                (index, json.dumps(data, ensure_ascii=False), "admin")
                for index, data in enumerate(records, start=30)
            ],
        )
        database.commit()
        database.close()

        response = self.client.get("/panel")
        self.assertEqual(response.status_code, 200)
        self.assertIn(b'id="yearTrendChart"', response.data)
        self.assertIn(b'id="statusDistributionChart"', response.data)
        self.assertIn("10.000,00 TL".encode(), response.data)
        self.assertIn("750,00 TL".encode(), response.data)
        self.assertIn('"status_counts": [0, 1, 2, 1]'.encode(), response.data)

    def test_login_is_generic_accessible_and_uses_semantic_messages(self):
        response = self.client.get("/logout", follow_redirects=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("Başarıyla çıkış yaptınız.".encode(), response.data)
        self.assertIn(b"auth-message-success", response.data)
        self.assertIn(
            'placeholder="Kullanıcı adı veya e-posta"'.encode(),
            response.data,
        )
        self.assertNotIn(b"admin@example.local", response.data)
        self.assertIn(b"data-password-toggle", response.data)
        self.assertIn(b'autocomplete="current-password"', response.data)
        self.assertIn(b"js/main.js", response.data)
        self.assertIn(b"favicon-32x32.png", response.data)

        invalid = self.client.post(
            "/login",
            data={"ident": "bilinmeyen", "password": "yanlis"},
        )
        self.assertEqual(invalid.status_code, 200)
        self.assertIn(b"auth-message-danger", invalid.data)

        favicon = self.client.get("/static/img/favicon.ico")
        self.assertEqual(favicon.status_code, 200)
        self.assertEqual(favicon.mimetype, "image/vnd.microsoft.icon")
        favicon.close()

    def test_password_forms_have_accessible_visibility_toggles(self):
        personnel = self.client.get("/personel")
        self.assertEqual(personnel.status_code, 200)
        self.assertIn(b'id="new-user-password"', personnel.data)
        self.assertIn(b'aria-controls="new-user-password"', personnel.data)
        self.assertIn(b'id="edit-password-1"', personnel.data)
        self.assertIn(b'aria-controls="edit-password-1"', personnel.data)

        password_change = self.client.get("/sifre-degistir")
        self.assertEqual(password_change.status_code, 200)
        for input_id in (
            "current-password",
            "change-password",
            "change-password-confirmation",
        ):
            self.assertIn(f'id="{input_id}"'.encode(), password_change.data)
            self.assertIn(
                f'aria-controls="{input_id}"'.encode(),
                password_change.data,
            )

        profile = self.client.get("/profil")
        self.assertEqual(profile.status_code, 200)
        for input_id in ("profile-password", "profile-password-confirmation"):
            self.assertIn(f'id="{input_id}"'.encode(), profile.data)
            self.assertIn(f'aria-controls="{input_id}"'.encode(), profile.data)

        database = sqlite3.connect(self.module.DB_PATH)
        admin_email = database.execute(
            "SELECT email FROM users WHERE id=1"
        ).fetchone()[0]
        database.close()
        token = self.module.serializer.dumps(admin_email, salt="sifre-sifirla")
        reset = self.client.get(f"/sifre-sifirla/{token}")
        self.assertEqual(reset.status_code, 200)
        for input_id in ("reset-password", "reset-password-confirmation"):
            self.assertIn(f'id="{input_id}"'.encode(), reset.data)
            self.assertIn(f'aria-controls="{input_id}"'.encode(), reset.data)

        self.assertIn('aria-label="Şifreyi göster"'.encode(), personnel.data)
        self.assertIn(b'autocomplete="new-password"', personnel.data)

    def test_form_schema_matches_real_a_to_ab_columns(self):
        database = sqlite3.connect(self.module.DB_PATH)
        database.row_factory = sqlite3.Row
        fields = database.execute(
            "SELECT field_key, label, is_system, position FROM form_fields ORDER BY position, id"
        ).fetchall()
        database.close()

        official = [row["field_key"] for row in fields]
        self.assertEqual(official, list(self.module.REAL_EXCEL_COLUMNS))
        self.assertEqual(len(official), 28)
        self.assertNotIn("kaynak_kalan_sure", official)
        labels = {row["field_key"]: row["label"] for row in fields}
        system_flags = {row["field_key"]: row["is_system"] for row in fields}
        self.assertEqual(labels["defter_sira_no"], "Sorumlu Personel")
        self.assertEqual(system_flags["defter_sira_no"], 1)
        self.assertNotIn("personel_adi", labels)

        create_form = self.client.get("/sirket-ekle")
        self.assertEqual(create_form.status_code, 200)
        self.assertIn("Sorumlu personel".encode(), create_form.data)
        self.assertNotIn(b'name="defter_sira_no"', create_form.data)
        self.assertNotIn(">Defter Sıra No<".encode(), create_form.data)
        self.assertIn(b"data-instrument-editor", create_form.data)
        self.assertIn("Başka ölçü aleti ekle".encode(), create_form.data)
        self.assertIn(b'name="attachments" type="file" multiple', create_form.data)
        self.assertIn(b"data-selected-files", create_form.data)
        self.assertIn("birden fazla dosya".encode(), create_form.data)

        response = self.client.get("/form-duzenle")
        self.assertEqual(response.status_code, 200)
        self.assertIn("28 / 28 eşleşti".encode(), response.data)
        self.assertNotIn("Sistem Meta Verisi".encode(), response.data)

        database = sqlite3.connect(self.module.DB_PATH)
        field_id = database.execute(
            "SELECT id FROM form_fields WHERE field_key='defter_sira_no'"
        ).fetchone()[0]
        database.close()
        response = self.client.post(f"/form-duzenle/{field_id}/sil")
        self.assertEqual(response.status_code, 302)
        database = sqlite3.connect(self.module.DB_PATH)
        still_exists = database.execute(
            "SELECT 1 FROM form_fields WHERE field_key='defter_sira_no'"
        ).fetchone()
        database.close()
        self.assertIsNotNone(still_exists)

    def test_multiple_instrument_pairs_are_saved_and_exported_together(self):
        form = self.valid_record_form()
        form["olcu_aleti_cinsi"] = ["Taksimetre", "Tanker sayacı"]
        form["olcu_aleti_sayisi"] = ["3", "2"]

        response = self.client.post("/sirket-ekle", data=form)
        self.assertEqual(response.status_code, 302)

        database = sqlite3.connect(self.module.DB_PATH)
        database.row_factory = sqlite3.Row
        company = database.execute(
            "SELECT * FROM companies ORDER BY sira_no DESC LIMIT 1"
        ).fetchone()
        database.close()
        company_data = json.loads(company["data"])
        self.assertEqual(
            company_data["olcu_aletleri"],
            [
                {"cinsi": "Taksimetre", "sayisi": 3},
                {"cinsi": "Tanker sayacı", "sayisi": 2},
            ],
        )
        self.assertEqual(company_data["olcu_aleti_sayisi"], "5")
        self.assertEqual(
            company_data["olcu_aleti_cinsi"],
            "Taksimetre (3); Tanker sayacı (2)",
        )

        detail = self.client.get(f"/sirket-oku?sira_no={company['sira_no']}")
        self.assertEqual(detail.status_code, 200)
        self.assertIn("Taksimetre (3); Tanker sayacı (2)".encode(), detail.data)

        exported = self.client.get("/disa-aktar/excel")
        self.assertEqual(exported.status_code, 200)
        workbook = load_workbook(io.BytesIO(exported.data), data_only=True)
        sheet = workbook.active
        self.assertEqual(sheet["F4"].value, "5")
        self.assertEqual(
            sheet["G4"].value,
            "Taksimetre (3); Tanker sayacı (2)",
        )
        self.assertIn("KOCAELİ İL MÜDÜRLÜĞÜ", sheet["A1"].value)
        self.assertIn("İdari Para Cezaları", sheet["A1"].value)
        self.assertIn("Dönem:", sheet["A2"].value)
        self.assertIn("Kayıt Sayısı: 1", sheet["A2"].value)
        self.assertIn("Toplam Ceza Tutarı:", sheet["A2"].value)
        self.assertIn("A1:F1", {str(item) for item in sheet.merged_cells.ranges})
        self.assertIn("A2:F2", {str(item) for item in sheet.merged_cells.ranges})
        self.assertGreaterEqual(sheet.row_dimensions[3].height, 60)
        self.assertEqual(sheet.freeze_panes, "A4")
        self.assertEqual(sheet.auto_filter.ref, "A3:AB4")
        self.assertFalse(sheet.sheet_view.showGridLines)
        self.assertEqual(sheet.page_setup.orientation, "landscape")
        self.assertTrue(sheet["A3"].alignment.wrap_text)
        self.assertEqual(sheet.column_dimensions["A"].width, 11)
        self.assertEqual(sheet.column_dimensions["G"].width, 42)
        workbook.close()

    def test_admin_can_create_view_download_and_delete_record_attachments(self):
        form = self.valid_record_form()
        form["defter_sira_no"] = "Formdan gönderilen sahte personel"
        form["attachments"] = [
            (io.BytesIO(b"%PDF-1.4 test attachment"), "tutanak.pdf"),
            (io.BytesIO(b"photo attachment"), "kanit.jpg"),
        ]
        response = self.client.post(
            "/sirket-ekle", data=form, content_type="multipart/form-data"
        )
        self.assertEqual(response.status_code, 302)

        database = sqlite3.connect(self.module.DB_PATH)
        database.row_factory = sqlite3.Row
        company = database.execute(
            "SELECT sira_no,data FROM companies ORDER BY sira_no DESC LIMIT 1"
        ).fetchone()
        company_data = json.loads(company["data"])
        attachments = database.execute(
            "SELECT * FROM attachments WHERE company_sira_no=? ORDER BY id",
            (company["sira_no"],),
        ).fetchall()
        database.close()
        self.assertEqual(company_data["defter_sira_no"], "admin")
        self.assertNotIn("personel_adi", company_data)
        self.assertEqual(len(attachments), 2)
        stored_paths = [
            Path(self.module.UPLOAD_DIR) / attachment["stored_name"]
            for attachment in attachments
        ]
        self.assertTrue(all(path.is_file() for path in stored_paths))

        view = self.client.get(f"/sirket-oku?sira_no={company['sira_no']}")
        self.assertEqual(view.status_code, 200)
        self.assertIn("tutanak.pdf".encode(), view.data)
        self.assertIn("kanit.jpg".encode(), view.data)
        download = self.client.get(f"/ek/{attachments[0]['id']}")
        self.assertEqual(download.status_code, 200)
        self.assertIn("attachment", download.headers["Content-Disposition"])
        self.assertEqual(download.data, b"%PDF-1.4 test attachment")
        download.close()

        deleted = self.client.post(
            "/sirket-sil", data={"sira_no": str(company["sira_no"])}
        )
        self.assertEqual(deleted.status_code, 302)
        database = sqlite3.connect(self.module.DB_PATH)
        self.assertEqual(
            database.execute("SELECT COUNT(*) FROM attachments").fetchone()[0], 0
        )
        database.close()
        self.assertTrue(all(not path.exists() for path in stored_paths))

    def test_attachment_limit_error_does_not_create_a_record(self):
        form = self.valid_record_form()
        form["attachments"] = [
            (io.BytesIO(f"file-{index}".encode()), f"belge-{index}.txt")
            for index in range(11)
        ]

        response = self.client.post(
            "/sirket-ekle", data=form, content_type="multipart/form-data"
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn("en fazla 10 dosya".encode(), response.data)
        database = sqlite3.connect(self.module.DB_PATH)
        company_count = database.execute(
            "SELECT COUNT(*) FROM companies"
        ).fetchone()[0]
        attachment_count = database.execute(
            "SELECT COUNT(*) FROM attachments"
        ).fetchone()[0]
        database.close()
        self.assertEqual(company_count, 0)
        self.assertEqual(attachment_count, 0)

    def test_personnel_attachment_moves_to_company_after_approval(self):
        database = sqlite3.connect(self.module.DB_PATH)
        cursor = database.execute(
            """INSERT INTO users(name,email,password_hash,role)
               VALUES('personel','personel@example.test','x','personel')"""
        )
        personnel_id = cursor.lastrowid
        database.commit()
        database.close()
        with self.client.session_transaction() as session:
            session["user_id"] = personnel_id
            session["user_name"] = "personel"
            session["role"] = "personel"

        form = self.valid_record_form()
        form["attachments"] = (io.BytesIO(b"video bytes"), "tespit.mp4")
        response = self.client.post(
            "/sirket-ekle", data=form, content_type="multipart/form-data"
        )
        self.assertEqual(response.status_code, 302)

        database = sqlite3.connect(self.module.DB_PATH)
        database.row_factory = sqlite3.Row
        pending = database.execute(
            "SELECT * FROM pending_actions WHERE status='bekliyor'"
        ).fetchone()
        pending_data = json.loads(pending["data"])
        pending_attachment = database.execute(
            "SELECT * FROM attachments WHERE pending_action_id=?", (pending["id"],)
        ).fetchone()
        database.close()
        self.assertEqual(pending_data["defter_sira_no"], "personel")
        self.assertNotIn("personel_adi", pending_data)
        self.assertIsNotNone(pending_attachment)
        pending_download = self.client.get(f"/ek/{pending_attachment['id']}")
        self.assertEqual(pending_download.status_code, 200)
        pending_download.close()

        with self.client.session_transaction() as session:
            session["user_id"] = 1
            session["user_name"] = "admin"
            session["role"] = "admin"
        approved = self.client.post(f"/onaylar/{pending['id']}/onayla")
        self.assertEqual(approved.status_code, 302)

        database = sqlite3.connect(self.module.DB_PATH)
        database.row_factory = sqlite3.Row
        moved = database.execute(
            "SELECT * FROM attachments WHERE id=?", (pending_attachment["id"],)
        ).fetchone()
        company_count = database.execute("SELECT COUNT(*) FROM companies").fetchone()[0]
        database.close()
        self.assertEqual(company_count, 1)
        self.assertIsNotNone(moved["company_sira_no"])
        self.assertIsNone(moved["pending_action_id"])

    def test_rejected_personnel_attachment_is_removed_from_database_and_disk(self):
        database = sqlite3.connect(self.module.DB_PATH)
        cursor = database.execute(
            """INSERT INTO users(name,email,password_hash,role)
               VALUES('personel','personel@example.test','x','personel')"""
        )
        personnel_id = cursor.lastrowid
        database.commit()
        database.close()
        with self.client.session_transaction() as session:
            session["user_id"] = personnel_id
            session["user_name"] = "personel"
            session["role"] = "personel"

        form = self.valid_record_form()
        form["attachments"] = (io.BytesIO(b"photo bytes"), "kanit.jpg")
        self.client.post(
            "/sirket-ekle", data=form, content_type="multipart/form-data"
        )
        database = sqlite3.connect(self.module.DB_PATH)
        database.row_factory = sqlite3.Row
        pending = database.execute(
            "SELECT * FROM pending_actions WHERE status='bekliyor'"
        ).fetchone()
        attachment = database.execute(
            "SELECT * FROM attachments WHERE pending_action_id=?", (pending["id"],)
        ).fetchone()
        database.close()
        stored_path = Path(self.module.UPLOAD_DIR) / attachment["stored_name"]
        self.assertTrue(stored_path.exists())

        with self.client.session_transaction() as session:
            session["user_id"] = 1
            session["user_name"] = "admin"
            session["role"] = "admin"
        missing_reason = self.client.post(f"/onaylar/{pending['id']}/reddet")
        self.assertEqual(missing_reason.status_code, 302)
        database = sqlite3.connect(self.module.DB_PATH)
        self.assertEqual(
            database.execute(
                "SELECT status FROM pending_actions WHERE id=?", (pending["id"],)
            ).fetchone()[0],
            "bekliyor",
        )
        self.assertEqual(
            database.execute(
                "SELECT COUNT(*) FROM attachments WHERE id=?", (attachment["id"],)
            ).fetchone()[0],
            1,
        )
        database.close()

        rejected = self.client.post(
            f"/onaylar/{pending['id']}/reddet",
            data={"decision_reason": "Belge okunamadığı için yeniden yüklenmeli."},
        )
        self.assertEqual(rejected.status_code, 302)
        database = sqlite3.connect(self.module.DB_PATH)
        attachment_count = database.execute(
            "SELECT COUNT(*) FROM attachments WHERE id=?", (attachment["id"],)
        ).fetchone()[0]
        company_count = database.execute("SELECT COUNT(*) FROM companies").fetchone()[0]
        decision = database.execute(
            "SELECT status,decision_reason FROM pending_actions WHERE id=?",
            (pending["id"],),
        ).fetchone()
        database.close()
        self.assertEqual(attachment_count, 0)
        self.assertEqual(company_count, 0)
        self.assertEqual(decision[0], "reddedildi")
        self.assertEqual(
            decision[1], "Belge okunamadığı için yeniden yüklenmeli."
        )
        self.assertFalse(stored_path.exists())


if __name__ == "__main__":
    unittest.main()
